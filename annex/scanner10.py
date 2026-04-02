import re
from datetime import datetime
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl import load_workbook
import numpy as np


"""
cell.coordinate	Full Excel reference	'C7'
cell.row	Row number (1-based)	7
cell.column	Column number (1-based)	3
cell.column_letter	Column letter ('A', 'B', ...)	'C'




cell in ws as a function of its coordinates

value = ws.cell(row=ROW_NUMBER, column=COLUMN_NUMBER).value
"""

COL_desde = 2
COL_hasta = 3
COL_desnivel_ida = 4
COL_desnivel_vuelta = 6
COL_desnivel_promedio = 4
COL_error = 8
COL_cota = 9



def format_float(value, decimals=3):
    try:
        f = float(value)
        return f"{round(f, decimals):.{decimals}f}"
    except (ValueError, TypeError):
        return ""

def str_to_float(str) :
    try:
        return np.round(float(str),3)
    except:
        return 0.000

class Block : 

    def __init__ ( 
            self,  start_point = "", start_z = "",
            end_point = "", end_z = "", go_level = "",
            come_level = "", mean_level = "", delta = "") : 

            self.start_point = start_point 
            self.end_point   = end_point 
            self.start_z     = str_to_float(start_z)
            self.end_z       = str_to_float(end_z )
            self.go_level    = str_to_float(go_level )
            self.come_level  = str_to_float(come_level )
            self.mean_level  = str_to_float(mean_level )
            self.delta       = str_to_float(delta)


    def __str__ (self) :

        return f"""
start_point = {self.start_point}
start_z = {self.start_z}
end_point = {self.end_point}
end_z = {self.end_z}
go_level = {self.go_level}
come_level = {self.come_level}
mean_level = {self.mean_level}
delta = {self.delta}
"""

class Scanner : 

    def __init__ (self, ws) :
        self.ws = ws
        self.blocks = []
        self.scan()
        

    

    def scan_block (self,ws,start_row) :

        START_ROW = start_row

        
        start_point = ""
        start_z = "",
        end_point = ""
        end_z = ""
        go_level = ""
        come_level = ""
        mean_level = ""
        delta = ""


        for row in ws.iter_rows(min_col = 2, max_col = 2, min_row = START_ROW):
            break_outer = False
            for cell in row:
                if cell.value is not None and  str(cell.value) != "":
                    start_point = cell.value
                    start_z = format_float(ws.cell(row=cell.row, column=9).value)
                    break_outer = True
                    break
            if break_outer:
                break

        for row in ws.iter_rows(min_col = 3, max_col = 3, min_row = START_ROW):
            break_outer = False
            for cell in row:
                if cell.value is not None and  str(cell.value) != "":
                    end_point = cell.value
                    end_z = format_float(ws.cell(row=cell.row, column=9).value  )
                    break_outer = True
                    break
            if break_outer:
                break
        

        for row in ws.iter_rows(min_col = 2, max_col = 2, min_row = START_ROW):
            break_outer = False
            for cell in row:
                if cell.value is not None and  str(cell.value) == "Desnivel":
                    go_level = format_float(ws.cell(row=cell.row, column=4).value)
                    come_level = format_float(ws.cell(row=cell.row, column=6).value)
                    delta = format_float(ws.cell(row=cell.row, column = 8).value)
                    mean_level = format_float(ws.cell(row=cell.row+1,column = 4).value)
                    break_outer = True
                    break
            if break_outer:
                break
        
        block = Block(
            start_point =start_point,
            start_z =start_z,
            end_point =end_point,
            end_z =end_z,
            go_level =go_level,
            come_level =come_level,
            mean_level =mean_level,
            delta =delta)
        
        self.blocks.append(block)

    def scan(self):
        ws = self.ws
        for row in ws.iter_rows(min_col = 2, max_col = 2, min_row = 1):
            for cell in row:
                if cell.value is not None and str(cell.value) == "DESDE":
                    self.scan_block(ws,cell.row+1)


        


def main () :
    input_file = "/home/jstvns/Q125/a10.xlsx"
    wb = load_workbook(input_file)
    sheet_names = wb.sheetnames
    ws  = wb[sheet_names[1]]
    scanner = Scanner(ws)
    scanner.scan()