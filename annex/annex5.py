

import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
import sys
import reader as rd
import re
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
import annexUtils
from openpyxl import load_workbook
import os
import glob
import annexImg

OFFSET   = 38
PAGEBREAKS = []

def generate (input_file='anexos/anteproyecto/anexo1.xlsx',output_file="test5.xlsx", src_dir="img", src_dir2="img_geo",crop_zoom=1.5) :
    
    print(f'\n\nGeneración de {output_file} en curso ...')
    
    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet("2.303.104.A (STC)")
    writer = Writer(workbook,worksheet)
    wb = load_workbook(input_file)
    ws = wb.active
    scanner = annexUtils.Scanner(ws)
    dst_dir = None
    dst_dir2 = None
    
    COL_WIDTHS = [
        0.10, #A 
        0.30, #B
        0.70, #C
        0.25, #D
        0.39, #E
        0.18, #F
        0.19, #G
        0.19, #H
        0.25, #I
        0.22, #J
        0.30, #K
        0.25, #L
        0.15, #M
        0.42, #N
        0.15, #O
        0.20, #P
        0.20, #Q
        0.33, #R
        0.25, #S
        0.15, #T
        0.25, #U
        0.15, #V
        0.33, #W
        0.28, #X
        0.46, #Y
        0.25, #Z
        0.10, #AA
    ]
    
    ROW_DICT = {
        0 :0.1,  
        6 :0.1,  
        7 :0.18, 
        8 :0.18, 
        9 :0.19, 
        10:0.1,
        12:0.12,
        13:0.12
        
    }
    
    #worksheet.autofit()
    annexUtils.set_column(worksheet,COL_WIDTHS)
    
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    
    # FIXED CONTENT
    writer.merge(f"B2:F6","",Format.BORDER)
    writer.merge(f"G2:Z3","FORMULARIO DE UBICACIÓN DE VÉRTICES DEL STC",
                 {"top":1, "right":1,"font_size":12,"bold":True,"align":"center","valign":"vcenter"})
    
    writer.merge(f"G4:Z6","FORMULARIO N°2.303.104.A",
                 {"bottom":1,"right":1, "font_size":12,"bold":True,"align":"center","valign":"vcenter"})
    
    writer.write(f"AA2","",Format.BLEFT)
    
    writer.merge(f"B7:Z7","",{"bottom":1})
    writer.merge(f"A8:A12","",{"right":1})
    writer.merge(f"AA8:AA12","",{"left":1})
    writer.merge(f"B13:Z13", "",{"top":1})
    
    writer.merge(f"B8:C8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B9:C9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B10:C10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B12:C12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"U12:Z12",f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
    
    writer.merge(f'D8:Y8', scanner.PROYECTO, Format.LEFT,Format.SIZE(10))
    writer.merge(f'D9:Y9',scanner.SECTOR,Format.SIZE(10))
    writer.merge(f'D10:Y10', scanner.TRAMO,Format.SIZE(10))
    writer.merge(f'D12:R12',scanner.REALIZADO,Format.SIZE(10))
    
    
  
    t_rows = scanner.get_t_rows()
    i = 0
    
    src_dir = None
        
    if src_dir2:
        dst_dir2 = annexImg.annex5_process_geo(src_dir2,crop_zoom=crop_zoom)
    
    
    # LOOP THE FOLLOWING CELLS 
    for r in t_rows :
        
        #CELL_nombre = ws[f'C{r}'].value # DONE
        POINT = scanner.get_est(r)
        
        CELL_f      = scanner.get_geo_s(r)
        CELL_l      = scanner.get_geo_w(r)
        CELL_h      = scanner.get_elip(r)
        
        POLY_NUM    = scanner.get_poligonal_num(r)
        
        CELL_X      = scanner.get_geo_x(r)
        CELL_Y      = scanner.get_geo_y(r)
        CELL_Z      = scanner.get_geo_z(r)
        
        CELL_N      = scanner.get_utm_n(r)
        CELL_E      = scanner.get_utm_e(r)
        
        CELL_altura = scanner.get_cota_orto(r)
        CELL_cota   = scanner.get_cota_geo(r)
        
        CELL_dm     = scanner.get_dm(r)
        CELL_dist   = scanner.get_dist(r)
        
        
        # CELL_NL     = ws[f'K{r}'].value
        # CELL_EL     = ws[f'L{r}'].value
        # CELL_MCL    = ws['H9'].value
        # CELL_Ko     = ws['H12'].value
        
        
        
        writer.write(f"B{15 + i * OFFSET}","Identificación del Vértice",Format.BOLD, Format.ITALIC, Format.SIZE(11))
        writer.write(f"H{15 + i * OFFSET}", "Nombre:", Format.SIZE(10))
        writer.write(f"O{15 + i * OFFSET}", "Dm. Ref.:",Format.SIZE(10))
        
        writer.write(
            f'F{16 + i * OFFSET}',
            "Pertenece a Poligonal Nº:",
            Format.SIZE(10)
        )
        
        writer.merge(
            f'M{16 + i * OFFSET}:N{16 + i * OFFSET}',
            POLY_NUM,
            Format.BBOTTOM, Format.CENTER, Format.SIZE(11)
        )
        
        writer.write(
            f'P{16 + i * OFFSET}',
            "Tipo de Poligonal (Ppal/Aux):",
            Format.SIZE(10)
        )
        
        writer.merge(
            f'X{16 + i * OFFSET}:Z{16 + i * OFFSET}',
            "PRINCIPAL",
            Format.BBOTTOM,Format.SIZE(11), Format.CENTER
        )
        
        writer.merge(f"K{15 + i * OFFSET}:M{15 + i * OFFSET}",POINT,Format.SIZE(11), Format.BOTTOM, Format.CENTER)
        writer.merge(f"N{25 + i * OFFSET}:P{25 + i * OFFSET}",CELL_altura,Format.NUM)
        writer.merge(f"X{25 + i * OFFSET}:Y{25 + i * OFFSET}",CELL_cota,Format.NUM)
        
        # UTM coordinates
        writer.merge(f'O{20 + i * OFFSET}:T{20 + i * OFFSET}',scanner.ZONA,Format.CENTER)
        writer.merge(f'O{21 + i * OFFSET}:T{21 + i * OFFSET}',scanner.MC, Format.CENTER)
        writer.merge(f"O{22 + i * OFFSET}:T{22 + i * OFFSET}",CELL_N,Format.NUM)
        writer.merge(f"O{23 + i * OFFSET}:T{23 + i * OFFSET}",CELL_E,Format.NUM)
        writer.write(f"U{22 + i * OFFSET}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"U{23 + i * OFFSET}","m",Format.SIZE(11),Format.LEFT)

        
        for k in range(len(scanner.PTL_N)):
            LTM_N = ws.cell(column=scanner.PTL_N[k],row=r).value
            LTM_E = ws.cell(column=scanner.PTL_E[k],row=r).value
            MCL   = scanner.MERIDIANO_CENTRAL[k]
            FACTOR = scanner.FACTOR_ESCALA[k]
            try:
                float(LTM_N)
            except:
                continue
            writer.merge(f"D{18 + i * OFFSET}:H{19 + i * OFFSET}", f"PTL{k+1}",Format.SIZE(11),Format.CENTER,Format.VCENTER)
            writer.merge(f"D{22 + i * OFFSET}:H{22 + i * OFFSET}",LTM_N,Format.RIGHT, Format.NUM)
            writer.merge(f"D{23 + i * OFFSET}:H{23 + i * OFFSET}",LTM_E,Format.RIGHT, Format.NUM)
            writer.merge(f"D{20 + i * OFFSET}:H{20 + i * OFFSET}",MCL,Format.LEFT, Format.CENTER)
            writer.merge(f"D{21 + i * OFFSET}:H{21 + i * OFFSET}",FACTOR, Format.DEC)
            break
        
        writer.merge(f"R{15 + i * OFFSET}:U{15 + i * OFFSET}", CELL_dm,Format.CENTER,Format.SIZE(10),Format.BOTTOM,Format.NUM2)
        writer.write(f"W{15 + i * OFFSET}","FECHA:",Format.SIZE(10))
        writer.merge(f"Y{15 + i * OFFSET}:Z{15 + i * OFFSET}", annexUtils.curr_date(1),Format.BOTTOM,Format.CENTER,Format.SIZE(10))
        
        writer.merge(f"N{18 + i * OFFSET}:S{19 + i * OFFSET}", "UTM",Format.SIZE(11),Format.CENTER,Format.VCENTER)
        
        
        writer.write(f"Q{25 + i * OFFSET}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"Z{25 + i * OFFSET}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"I{22 + i * OFFSET}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"I{23 + i * OFFSET}","m",Format.SIZE(11),Format.LEFT)
        
        writer.merge(f'B{17 + i * OFFSET}:Z{17 + i * OFFSET}' , 'Coordenadas', Format.SIZE(11), Format.CENTER)
        writer.write(f"N{20 + i * OFFSET}","Huso:",Format.SIZE(11),Format.LEFT)
        writer.write(f"N{21 + i * OFFSET}","MC:",Format.SIZE(11),Format.LEFT)
        writer.write(f"N{22 + i * OFFSET}","N:",Format.SIZE(11),Format.LEFT)
        writer.write(f"N{23 + i * OFFSET}","E:",Format.SIZE(11),Format.LEFT)
        
        writer.write(f"C{20 + i * OFFSET}","MCL:",Format.SIZE(11),Format.LEFT)
        writer.write(f"C{21 + i * OFFSET}","Ko:",Format.SIZE(11),Format.LEFT)
        writer.write(f"C{22 + i * OFFSET}","NL:",Format.SIZE(11),Format.LEFT)
        writer.write(f"C{23 + i * OFFSET}","EL:",Format.SIZE(11),Format.LEFT)
        
        writer.write(f"F{25 + i * OFFSET}","Altura (n.m.m. modelada):",Format.SIZE(11))
        writer.write(f"S{25 + i * OFFSET}","Cota (nivelada):",Format.SIZE(11))
        
        
        
        
        
        if False:
            cleaned_point = POINT.replace("-", "")
            
            img_path_a = os.path.join(dst_dir, f'{cleaned_point}*', f'{cleaned_point}_a.*')
            img_path_p = os.path.join(dst_dir, f'{cleaned_point}*', f'{cleaned_point}_p.*')
            
            match_a = glob.glob(img_path_a)
            match_p = glob.glob(img_path_p)
            
            if match_a:
                worksheet.insert_image(f'B{41 + i * OFFSET}', match_a[0] , {'object_position': 1})
            else :
                writer.merge(
                    f"B{41 + i * OFFSET}:F{48 + i * OFFSET}","Fotografía\nDetalle",
                    Format.BORDER,
                    Format.CENTER,Format.VCENTER)
            
            if match_p:
                worksheet.insert_image(f'B{27 + i * OFFSET}', match_p[0],  {'object_position': 1})
            else:
                writer.merge(
                    f"B{27 + i * OFFSET}:L{39 + i * OFFSET}","Fotografía\nPanorámica",
                    Format.BORDER,
                    Format.CENTER,Format.VCENTER)
        
     
        if src_dir2 :

            cleaned_point = POINT.replace("-", "")

            img_path_a = os.path.join(dst_dir2, f'{cleaned_point}_a.*')
            img_path_p = os.path.join(dst_dir2, f'{cleaned_point}_p.*')
            img_path_g = os.path.join(dst_dir2, f'{cleaned_point}_g.*')
            
            match_a = glob.glob(img_path_a)
            match_p = glob.glob(img_path_p)
            match_g = glob.glob(img_path_g)
            
            if match_a:
                worksheet.insert_image(f'B{41 + i * OFFSET}', match_a[0] , {'object_position': 1})
            else :
                writer.merge(
                    f"B{41 + i * OFFSET}:F{48 + i * OFFSET}","Fotografía\nDetalle",
                    Format.BORDER,
                    Format.CENTER,Format.VCENTER)
            
            if match_p:
                worksheet.insert_image(f'B{27 + i * OFFSET}', match_p[0],  {'object_position': 1})
            else:
                writer.merge(
                    f"B{27 + i * OFFSET}:L{39 + i * OFFSET}","Fotografía\nPanorámica",
                    Format.BORDER,
                    Format.CENTER,Format.VCENTER)
                
            
            if match_g:
                worksheet.insert_image(f'N{27 + i * OFFSET}', match_g[0] , {'object_position': 1})
            else:
                writer.merge(
                    f"N{27 + i * OFFSET}:Z{39 + i * OFFSET}","Vista\nAérea",
                    Format.BORDER,
                    Format.CENTER,Format.VCENTER)
        
        
        writer.merge(f"H{41 + i * OFFSET}:Z{41 + i * OFFSET}","Descripción",Format.BOTTOM,Format.LEFT,Format.SIZE(10))
        
        writer.write(f'I{45 + i * OFFSET}', "Distancia a la:",Format.SIZE(9))
        writer.write(f'I{46 + i * OFFSET}', "ruta:",Format.SIZE(9))

        
        writer.write(f'N{45 + i * OFFSET}', f"{CELL_dist} m", Format.SIZE(9),Format.LEFT)
        
        writer.write(f'I{43 + i * OFFSET}', "Materialidad:",Format.SIZE(9))
        writer.write(f'N{43 + i * OFFSET}', "MONOLITO DE HORMIGÓN, PINTADO DE AMARILLO", Format.SIZE(9))

        writer.write(f'I{44 + i * OFFSET}', "Dimensiones:",Format.SIZE(9))
        writer.write(f'N{44 + i * OFFSET}', "D: 15 cm. FIERRO ESTRIADO 12 mm.", Format.SIZE(9))
        #writer.merge(f'L{44 + i * OFFSET}:L{44 + i * OFFSET}', "D: 15 cm. FIERRO ESTRIADO 12 mm.", Format.SIZE(9))
        
        writer.merge(f"G{42 + i * OFFSET}:G{48 + i * OFFSET}","",Format.BRIGHT)
        writer.merge(f"AA{42 + i * OFFSET}:AA{48 + i * OFFSET}","",Format.BLEFT)
        writer.merge(f"H{49 + i * OFFSET}:Z{49 + i * OFFSET}","",Format.TOP)
        writer.merge(f"A{50 + i * OFFSET}:AA{50 + i * OFFSET}","",{})
        
        ROW_DICT.update({15 + i * OFFSET : 0.25})
        ROW_DICT.update({16 + i * OFFSET : 0.45})
        ROW_DICT.update({17 + i * OFFSET : 0.14})
        ROW_DICT.update({18 + i * OFFSET : 0.14})
        
        ROW_DICT.update({key:0.16 for key in range(27 + i*OFFSET , 39 + i*OFFSET)})  
        ROW_DICT.update({key:0.175 for key in range(41 + i*OFFSET , 48 + i*OFFSET)})  
        
        PAGEBREAKS.append(50 + i * OFFSET)
        
        i += 1
    
    worksheet.set_h_pagebreaks(PAGEBREAKS)
    annexUtils.set_row_dict(worksheet,ROW_DICT)
    
    # formatter.set_rows({0:2,1:2,2:2, 4:2})
    workbook.close()
    
    print(f'\nGeneración de {output_file} completada\n')




if __name__ == "__main__":
    
    generate(src_dir = '', src_dir2= '')
    
    #f1 = sys.argv[1]
    #f2 = sys.argv[2]
    
    #reader = rd.Reader (f1, "", f2)
    #matrix, labels, om, ol, heights = reader.getData()
    #model = mdl.Model(heights,matrix,labels, om, ol)
    
    #trans (model)

