import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
import sys
import reader as rd
import re
import os
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
import annexUtils
import annexImg
import glob
from openpyxl import load_workbook


OFFSET   = 38
PAGEBREAKS = []

def generate (input_file='anexos/anteproyecto/anexo1.xlsx',output_file="test2.xlsx", src_dir="img", src_dir2="img_geo", crop_zoom=1.5) :
    
    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet("2.903.3.F (RRP)")
    writer = Writer(workbook,worksheet)
    wb = load_workbook(input_file)
    ws = wb.active
    scanner = annexUtils.Scanner(ws)
    
    COL_WIDTHS = [
        0.10, #A 
        0.25, #B
        0.75, #C
        0.25, #D
        0.35, #E
        0.18, #F
        0.19, #G
        0.19, #H
        0.25, #I
        0.22, #J
        0.40, #K
        0.20, #L
        0.07, #M
        0.35, #N
        0.15, #O
        0.25, #P
        0.28, #Q
        0.33, #R
        0.25, #S
        0.10, #T
        0.25, #U
        0.15, #V
        0.33, #W
        0.19, #X
        0.46, #Y
        0.25, #Z
        0.10, #AA
    ]
    
    ROW_DICT = {
        0 :0.1,  
        6 :0.1,  
        7 :0.18, 
        8 :0.18, 
        9 :0.19, 
        10:0.1,
        12:0.12,
        13:0.12
    }
    
    
    #worksheet.autofit()
    annexUtils.set_column(worksheet,COL_WIDTHS)
    
    
    # Page configuration
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)
    
    # FIXED CONTENT
    writer.merge(f"B2:F6","",Format.BORDER)
    writer.merge(f"G2:Z3","PUNTOS DE LA RED DE REFERENCIA PRINCIPAL",
                 {"top":1, "right":1,"font_size":12,"bold":True,"align":"center","valign":"vcenter"})
    
    writer.merge(f"G4:Z6","FORMULARIO N° 2.903.3.F",
                 {"bottom":1,"right":1, "font_size":12,"bold":True,"align":"center","valign":"vcenter"})
    
    writer.write(f"AA2","",Format.LEFT)
    
    writer.merge(f"B7:Z7","",{"bottom":1})
    writer.merge(f"A8:A12","",{"right":1})
    writer.merge(f"AA8:AA12","",{"left":1})
    writer.merge(f"B13:Z13", "",{"top":1})
    
    writer.merge(f"B8:C8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B9:C9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B10:C10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B12:C12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"U12:Z12",f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
    
    writer.merge(f'D8:Y8', scanner.PROYECTO,Format.SIZE(10))
    writer.merge(f'D9:Y9',scanner.SECTOR,Format.SIZE(10))
    writer.merge(f'D10:Y10', scanner.TRAMO,Format.SIZE(10))
    writer.merge(f'D12:R12',scanner.REALIZADO,Format.SIZE(10))
  
    if src_dir2:
        dst_dir2 = annexImg.annex2_process_geo(src_dir2,crop_zoom=crop_zoom)
        
  
    FRST_ROW = 0
    LST_ROW  = 0
    
    for merged_cell_range in ws.merged_cells.ranges:
        if merged_cell_range.min_col == 2 and merged_cell_range.max_col == 2:
            first_cell = ws.cell(row=merged_cell_range.min_row, column=2).value
            if first_cell == "RRP":
                FRST_ROW = merged_cell_range.min_row
                LST_ROW  = merged_cell_range.max_row
                first_row = merged_cell_range.min_row
                last_row = merged_cell_range.max_row
                break
 
    
    
    # LOOP THE FOLLOWING CELLS
    CORRECTION = 0
    for i,r in enumerate(range(FRST_ROW,LST_ROW+1)):
        
        CONST = i * OFFSET + CORRECTION
        
        POINT       = scanner.get_est(r)
        CELL_f      = scanner.get_geo_s(r)
        CELL_l      = scanner.get_geo_w(r)
        CELL_h      = scanner.get_elip(r)
        
        CELL_X      = scanner.get_geo_x(r)
        CELL_Y      = scanner.get_geo_y(r)
        CELL_Z      = scanner.get_geo_z(r)
        
        CELL_N      = scanner.get_utm_n(r)
        CELL_E      = scanner.get_utm_e(r)
        
        CELL_altura = scanner.get_cota_orto(r)
        CELL_cota   = scanner.get_cota_geo(r)

        CELL_dist   = scanner.get_dist(r)
        CELL_dm     = scanner.get_dm(r)
        
        writer.write(f"B{15 + CONST}","Identificación del Punto",Format.BOLD, Format.ITALIC, Format.SIZE(11))

        writer.write(f"H{15 + CONST}", "Nombre:", Format.SIZE(10))
        writer.merge(f"K{15 + CONST}:M{15 + CONST}",POINT,Format.SIZE(11), Format.BOTTOM, Format.CENTER)

        writer.write(f"O{15 + CONST}", "Dm. Ref.:",Format.SIZE(10))
        writer.merge(f"R{15 + CONST}:U{15 + CONST}", CELL_dm, Format.CENTER,Format.SIZE(10),Format.BOTTOM)
        
        
        writer.write(
            f"C{21 + CONST}",
            CELL_f,
        )
        
        writer.write(
            f"C{22 + CONST}",
            CELL_l,
        )
        
        writer.merge(
            f"C{23 + CONST}:D{23 + CONST}",
            CELL_h,
            Format.CENTER, Format.NUM
        )
        
        writer.merge(
            f"H{21 + CONST}:K{21 + CONST}",
            CELL_X,
            Format.RIGHT,Format.NUM
        )
        
        writer.merge(
            f"H{22 + CONST}:K{22 + CONST}",
            CELL_Y,
            Format.RIGHT,Format.NUM
        )
        
        writer.merge(
            f"H{23 + CONST}:K{23 + CONST}",
            CELL_Z,
            Format.RIGHT, Format.NUM
        )
        
        writer.merge(f'P{20 + CONST}:R{20 + CONST}',scanner.ZONA, Format.CENTER)
        writer.merge(f'P{21 + CONST}:R{21 + CONST}',scanner.MC, Format.CENTER)
        writer.merge(f"O{22 + CONST}:R{22 + CONST}",CELL_N,Format.RIGHT,Format.NUM)
        writer.merge(f"O{23 + CONST}:R{23 + CONST}",CELL_E,Format.RIGHT,Format.NUM)
        
        

        writer.write(f"W{15 + CONST}","FECHA:",Format.SIZE(10))
        writer.merge(f"Y{15 + CONST}:Z{15 + CONST}", annexUtils.curr_date(1),Format.BOTTOM,Format.CENTER,Format.SIZE(10))
        
        writer.merge(f"B{17 + CONST}:Z{17 + CONST}", "Coordenadas",Format.SIZE(11),Format.CENTER)
        writer.merge(f"B{18 + CONST}:F{18 + CONST}", "Geodésicas",Format.SIZE(11),Format.CENTER)
        writer.merge(f"B{19 + CONST}:F{19 + CONST}", "Ref. SIRGAS", Format.SIZE(11),Format.CENTER)
        
        writer.merge(f"G{18 + CONST}:L{19 + CONST}", "Geocéntricas",Format.SIZE(11), Format.CENTER, Format.VCENTER )
        writer.merge(f"N{18 + CONST}:S{19 + CONST}", "UTM",Format.SIZE(11),Format.CENTER,Format.VCENTER)
        
        
        writer.write(f"B{21 + CONST}","φ:",Format.SIZE(11),Format.LEFT)
        writer.write(f"B{22 + CONST}","λ:",Format.SIZE(11),Format.LEFT)
        writer.write(f"B{23 + CONST}","h:",Format.SIZE(10),Format.LEFT)
        
        writer.write(f"G{21 + CONST}","X:",Format.SIZE(11),Format.LEFT)
        writer.write(f"G{22 + CONST}","Y:",Format.SIZE(11),Format.LEFT)
        writer.write(f"G{23 + CONST}","Z:",Format.SIZE(11),Format.LEFT)
        
        
        writer.write(f"L{21 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"L{22 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"L{23 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"E{23 + CONST}","m",Format.SIZE(11),Format.LEFT)

        writer.write(f"S{22 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"S{23 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"Z{22 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"Z{23 + CONST}","m",Format.SIZE(11),Format.LEFT)
        
        
        
        writer.write(f"N{20 + CONST}","Huso:",Format.SIZE(11),Format.LEFT)
        writer.write(f"N{21 + CONST}","MC:",Format.SIZE(11),Format.LEFT)
        writer.write(f"N{22 + CONST}","N:",Format.SIZE(11),Format.LEFT)
        writer.write(f"N{23 + CONST}","E:",Format.SIZE(11),Format.LEFT)
        
        
        #############
        # ITERATION #
        #############
        
        PTL_OFFSET = 0
        
        for k in range(len(scanner.PTL_N)):
            LTM_N = ws.cell(column=scanner.PTL_N[k],row=r).value
            LTM_E = ws.cell(column=scanner.PTL_E[k],row=r).value
            MCL   = scanner.MERIDIANO_CENTRAL[k]
            FACTOR = scanner.FACTOR_ESCALA[k]
            
            try:
                float(LTM_N)
            except:
                continue
            
            writer.merge(f"U{18 + CONST + PTL_OFFSET }:Z{19 + CONST + PTL_OFFSET}", f"PTL-{k+1}",Format.SIZE(11),Format.CENTER,Format.VCENTER)
            writer.write(f"U{20 + CONST + PTL_OFFSET }","MCL:",Format.SIZE(11),Format.LEFT)
            writer.write(f"U{21 + CONST + PTL_OFFSET}","Ko:",Format.SIZE(11),Format.LEFT)
            writer.write(f"U{22 + CONST + PTL_OFFSET}","NL:",Format.SIZE(11),Format.LEFT)
            writer.write(f"U{23 + CONST + PTL_OFFSET}","EL:",Format.SIZE(11),Format.LEFT)
            writer.merge(f"W{20 + CONST + PTL_OFFSET}:Z{20 + CONST + PTL_OFFSET}", MCL,Format.LEFT)
            writer.merge(f"W{21 + CONST + PTL_OFFSET}:Z{21 + CONST + PTL_OFFSET}", FACTOR,Format.LEFT,Format.DEC)
            writer.merge(f"V{22 + CONST + PTL_OFFSET}:Y{22 + CONST + PTL_OFFSET}",LTM_N,Format.RIGHT,Format.NUM)
            writer.merge(f"V{23 + CONST + PTL_OFFSET}:Y{23 + CONST + PTL_OFFSET}",LTM_E,Format.RIGHT,Format.NUM)
            PTL_OFFSET += 6
        
        CONST_prev = CONST
        CORRECTION += PTL_OFFSET - 6
        CONST += PTL_OFFSET - 6
        
        #-------------------------------------------------------------------------------
        ###############
        # SUB 24 ROWS #
        ###############
        writer.write(f"E{25 + CONST}","Altura (n.m.m. modelada):",Format.SIZE(11))
        writer.write(f"R{25 + CONST}","Cota (nivelada):",Format.SIZE(11))
        writer.write(f"P{25 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"Z{25 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.merge(f"L{25 + CONST}:O{25 + CONST}",CELL_altura, Format.RIGHT, Format.NUM)
        writer.merge(f"W{25 + CONST}:Y{25 + CONST}",CELL_cota,Format.RIGHT, Format.NUM)
     
        if src_dir2 :
            cleaned_point = POINT.replace("-", "")

            img_path_g = os.path.join(dst_dir2, f'{cleaned_point}_g.*')
            img_path_a = os.path.join(dst_dir2, f'{cleaned_point}_a.*')
            img_path_p = os.path.join(dst_dir2, f'{cleaned_point}_p.*')

            match_g = glob.glob(img_path_g)
            match_a = glob.glob(img_path_a)
            match_p = glob.glob(img_path_p)
            

            if match_a:
                worksheet.insert_image(f'B{41 + CONST}', match_a[0] , {'object_position': 1})
            else:
                writer.merge(f"B{41 + CONST}:F{48 + CONST}","Fotografía\nDetalle",Format.BORDER,Format.CENTER,Format.VCENTER)
            
            if match_p:
                worksheet.insert_image(f'B{27 + CONST}', match_p[0],  {'object_position': 1})
            else:
                writer.merge(f"B{27 + CONST}:L{39 + CONST}","Fotografía\nPanorámica",Format.BORDER,Format.CENTER,Format.VCENTER)
        
            if match_g:
                worksheet.insert_image(f'N{27 + CONST}', match_g[0] , {'object_position': 1})
            else :
                writer.merge(f"N{27 + CONST}:Z{39 + CONST}","Vista\nAérea",Format.BORDER,Format.CENTER,Format.VCENTER)
        
        
        writer.merge(f"H{41 + CONST}:Z{41 + CONST}","Descripción",Format.BOTTOM,Format.LEFT,Format.SIZE(10))
        
        writer.write(f'I{43 + CONST}', "Materialidad:",Format.SIZE(9))
        writer.write(f'I{44 + CONST}', "Dimensiones:",Format.SIZE(9))

        writer.write(f'I{45 + CONST}', "Distancia a la:",Format.SIZE(9))
        writer.write(f'I{46 + CONST}', "ruta",Format.SIZE(9))

        
        writer.write(f'M{43 + CONST}', "MONOLITO DE HORMIGÓN, PINTADO DE AMARILLO", Format.SIZE(9))
        writer.write(f'M{44 + CONST}', "30,0 cm x 30,0 cm x 50,0 cm", Format.SIZE(9))
        writer.merge(f'M{45 + CONST}:P{45 + CONST}', f"{CELL_dist} m", Format.SIZE(9), Format.LEFT)


        writer.merge(f"G{42 + CONST}:G{48 + CONST}","",Format.BRIGHT)
        writer.merge(f"AA{42 + CONST}:AA{48 + CONST}","",Format.BLEFT)
        writer.merge(f"H{49 + CONST}:Z{49 + CONST}","",Format.TOP)
        writer.merge(f"A{50 + CONST}:AA{50 + CONST}","",{})
        
        ROW_DICT.update({key:0.16 for key in range(27 + CONST , 39 + CONST)})
        ROW_DICT.update({key:0.175 for key in range(41 + CONST , 48 + CONST)})  
        
        PAGEBREAKS.append(50 + CONST)

    
    worksheet.set_h_pagebreaks(PAGEBREAKS)
    annexUtils.set_row_dict(worksheet,ROW_DICT)
    
    # formatter.set_rows({0:2,1:2,2:2, 4:2})
    workbook.close()




if __name__ == "__main__":
    
    generate()
    #f1 = sys.argv[1]
    #f2 = sys.argv[2]
    
    #reader = rd.Reader (f1, "", f2)
    #matrix, labels, om, ol, heights = reader.getData()
    #model = mdl.Model(heights,matrix,labels, om, ol)
    
    #trans (model)

