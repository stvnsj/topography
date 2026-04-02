import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
import sys
import reader as rd
import re
import annexUtils
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
from annexUtils import Scanner
from openpyxl   import load_workbook


def generate (input_file = 'anexos/anteproyecto/annex1.xlsx' , output_file = "test4.xlsx") :
    
    workbook = xlsxwriter.Workbook(output_file)
    wb = load_workbook(input_file)
    ws = wb.active
    scanner = annexUtils.Scanner(ws)
    column = 'B'
    
    polygones = []
    
    min_row = 0
    max_row = 0
    
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_col == openpyxl.utils.column_index_from_string(column):
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            if top_left_cell.value == "RRP":
                min_row = merged_range.min_row
                max_row = merged_range.max_row
    
    COL_WIDTHS = [
        0.10, #A
        0.53, #B
        0.25, #C
        1.15, #D
        0.30, #E
        0.02, #F
        0.13, #G
        1.20, #H
        0.15, #I
        0.02, #J
        0.15, #K
        0.55, #L
        0.85, #M
        0.15, #N
        0.02, #O
        0.30, #P
        0.80, #Q
        0.75, #R
        0.15, #S
        0.02, #T
        0.10, #U
    ]
    
    HEIGHT_DICT = {0:0.12 , 6:0.12 , 7:0.18 , 8:0.18 , 9:0.18 , 10:0.08, 12:0.12}
 
    worksheet = workbook.add_worksheet("2.903.3.G")
    
    annexUtils.set_column(worksheet,COL_WIDTHS,INCH_COL=10.14)
    #annexUtils.set_row(worksheet,ROW_HEIGHTS)
    
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)
    
    writer = Writer(workbook,worksheet)
    
    # FIXED CONTENT
    writer.merge(f"B2:D6","",Format.BORDER)
    writer.merge(
        f"E2:T5","RESUMEN DE COORDENADAS DE LA RED DE REFERENCIA PRINCIPAL",
        Format.BTOP,Format.BRIGHT,Format.SIZE(12),Format.BOLD,
        Format.CENTER, Format.VCENTER
    )
    
    writer.merge(
        f"E6:T6",
        "FORMULARIO N°2.903.3.G",
        Format.BBOTTOM,Format.BRIGHT,Format.SIZE(12),
        Format.BOLD,Format.CENTER
    )
    
    
    writer.write("U2","",Format.BLEFT)
    writer.merge("B7:T7","",{"bottom":1})
    writer.merge("A8:A12","",{"right":1})
    writer.merge("U8:U12","",{"left":1})
    writer.merge("B13:T13", "",{"top":1, "bottom":1})
    
    writer.merge("B8:C8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge("B9:C9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge("B10:C10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge("B12:C12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge("N12:S12",f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
    
    writer.merge(f'D8:R8', scanner.PROYECTO,Format.SIZE(10))
    writer.merge(f'D9:R9',scanner.SECTOR,Format.SIZE(10))
    writer.merge(f'D10:R10', scanner.TRAMO,Format.SIZE(10))
    writer.merge(f'D12:M12',scanner.REALIZADO,Format.SIZE(10))
    
    
    writer.merge("B14:B16", 'Punto', Format.SIZE(9), Format.CENTER,Format.VCENTER,Format.BORDER)
    writer.write("D14", 'Geodésicas   Ref.', Format.SIZE(9), Format.LEFT)
    writer.write("D15", 'SIRGAS', Format.SIZE(9), Format.LEFT)
    writer.merge("C16:F16", '', Format.BBOTTOM, Format.BRIGHT)
    writer.merge("F14:F15", '', Format.BRIGHT)
    
    writer.merge("G15:I15",'Geocéntricas',Format.SIZE(9),Format.CENTER,Format.VCENTER)
    writer.merge("G16:J16",'',Format.BBOTTOM,Format.BRIGHT)
    writer.merge("J14:J15",'',Format.BRIGHT)
    
    writer.merge("K15:N15", 'UTM', Format.SIZE(9),Format.CENTER,Format.VCENTER)
    writer.merge("K16:O16", '', Format.BBOTTOM, Format.BRIGHT)
    writer.merge("O14:O15", '', Format.BRIGHT)
    
    writer.merge("P15:S15", 'LTM', Format.SIZE(9),Format.CENTER,Format.VCENTER)
    writer.merge("P16:T16", '', Format.BBOTTOM, Format.BRIGHT)
    writer.merge("T14:T15", '', Format.BRIGHT)
    
    curr_row = 17
    
    scanner = Scanner(ws)
    
    i = 0
    
    
    for row in range(min_row,max_row+1):
        
        
        # ESTACION
        POINT = scanner.get_est(row)
        
        # COORDENADAS GEODÉSICAS
        GEO_f = scanner.get_geo_s(row)
        GEO_l = scanner.get_geo_w(row)
        
        # ALTURA ELIP
        GEO_h = scanner.get_elip(row)
        
        # GEOCÉNTRICAS
        GEO_X = scanner.get_geo_x(row)
        GEO_Y = scanner.get_geo_y(row)
        GEO_Z = scanner.get_geo_z(row)
        
        # COORDENADAS UTM
        UTM_N = scanner.get_utm_n(row)
        UTM_E = scanner.get_utm_e(row)
        UTM_H = scanner.get_cota_orto(row)
        
        LTM_C = scanner.get_cota_geo(row)
        
        for k in range(len(scanner.PTL_N)):
            
            
            LTM_N = ws.cell(column=scanner.PTL_N[k],row=row).value
            LTM_E = ws.cell(column=scanner.PTL_E[k],row=row).value
            
            
            try:
                float(LTM_N)
            except:
                continue
            
            
            if len(scanner.PTL_N) == 1:
                writer.merge(
                    f"B{curr_row}:B{curr_row+3}",
                    POINT,
                    Format.CENTER,
                    Format.VCENTER,
                    Format.SIZE(10),
                    Format.BORDER,
                    Format.NUM
                )
            else:
                writer.merge(
                    f"B{curr_row}:B{curr_row+3}",
                    f'{POINT}\n(PTL-{k+1})',
                    Format.CENTER,
                    Format.VCENTER,
                    Format.SIZE(10),
                    Format.BORDER,
                    Format.NUM
                )
            
            writer.write(f'C{curr_row}','φ:'  ,Format.SIZE(10),Format.CENTER)
            writer.write(f'C{curr_row+1}','λ:',Format.SIZE(10),Format.CENTER)
            writer.write(f'C{curr_row+2}','h:',Format.SIZE(10),Format.CENTER)
            
            writer.merge(f'D{curr_row}:E{curr_row}',    GEO_f,Format.SIZE(10),Format.CENTER)
            writer.merge(f'D{curr_row+1}:E{curr_row+1}',GEO_l,Format.SIZE(10),Format.CENTER)
            writer.write(f'D{curr_row+2}',GEO_h,Format.SIZE(10),Format.RIGHT,Format.NUM)
            writer.write(f'E{curr_row+2}','m',Format.LEFT,Format.SIZE(10))
            
            writer.merge(f'F{curr_row}:F{curr_row+2}','',Format.BRIGHT)
            writer.merge(f'C{curr_row+3}:F{curr_row+3}','',Format.BBOTTOM,Format.BRIGHT)
            
            writer.write(f'G{curr_row}','X:',Format.SIZE(10)  ,Format.RIGHT)
            writer.write(f'G{curr_row+1}','Y:',Format.SIZE(10),Format.RIGHT)
            writer.write(f'G{curr_row+2}','Z:',Format.SIZE(10),Format.RIGHT)
            
            writer.write(f'H{curr_row}'  ,GEO_X,Format.SIZE(10),Format.RIGHT,Format.NUM)
            writer.write(f'H{curr_row+1}',GEO_Y,Format.SIZE(10),Format.RIGHT,Format.NUM)
            writer.write(f'H{curr_row+2}',GEO_Z,Format.SIZE(10),Format.RIGHT,Format.NUM)
            
            writer.write(f'I{curr_row}','m',Format.SIZE(10))
            writer.write(f'I{curr_row+1}','m',Format.SIZE(10))
            writer.write(f'I{curr_row+2}','m',Format.SIZE(10))
            
            writer.merge(f'J{curr_row}:J{curr_row+2}','',Format.BRIGHT)
            writer.merge(f'G{curr_row+3}:J{curr_row+3}','',Format.BBOTTOM,Format.BRIGHT)
            
            writer.write(f'K{curr_row}','N:',Format.SIZE(10))
            writer.write(f'K{curr_row+1}','E:',Format.SIZE(10))
            writer.merge(f'K{curr_row+2}:L{curr_row+2}','H (model):',Format.SIZE(10))
            
            writer.merge(f'L{curr_row}:M{curr_row}',UTM_N,Format.SIZE(10),Format.NUM)
            writer.merge(f'L{curr_row+1}:M{curr_row+1}',UTM_E,Format.SIZE(10), Format.NUM)
            writer.write(f'M{curr_row+2}',UTM_H,Format.SIZE(10),Format.NUM)
            
            writer.write(f'N{curr_row}','m',Format.SIZE(10))
            writer.write(f'N{curr_row+1}','m',Format.SIZE(10))
            writer.write(f'N{curr_row+2}','m',Format.SIZE(10))
            
            writer.merge(f'O{curr_row}:O{curr_row+2}','',Format.BRIGHT)
            writer.merge(f'K{curr_row+3}:O{curr_row+3}','',Format.BBOTTOM,Format.BRIGHT)
            
            writer.write(f'P{curr_row}','NL:',Format.SIZE(10))
            writer.write(f'P{curr_row+1}','EL:',Format.SIZE(10))
            writer.merge(f'P{curr_row+2}:Q{curr_row+2}','Cota (nivelada):',Format.SIZE(10))
            
            
            writer.merge(f'Q{curr_row}:R{curr_row}',LTM_N,Format.SIZE(10))
            writer.merge(f'Q{curr_row+1}:R{curr_row+1}',LTM_E,Format.SIZE(10))
            writer.write(f'R{curr_row+2}',LTM_C,Format.SIZE(10),Format.NUM)
            
            writer.write(f'S{curr_row}','m',Format.SIZE(10))
            writer.write(f'S{curr_row+1}','m',Format.SIZE(10))
            writer.write(f'S{curr_row+2}','m',Format.SIZE(10))
            
            writer.merge(f'T{curr_row}:T{curr_row+2}','',Format.BRIGHT)
            writer.merge(f'P{curr_row+3}:T{curr_row+3}','',Format.BBOTTOM,Format.BRIGHT)
            
            if i > 0:
                writer.merge(f'B{curr_row-1}:T{curr_row-1}','',Format.BORDER)
                HEIGHT_DICT[curr_row-2]=0.15
            curr_row += 5
            i += 1
        
        #curr_row -= 1
    writer.merge(f'B{curr_row-1}:T{curr_row-1}','',Format.BTOP)

    annexUtils.set_row_dict(worksheet,HEIGHT_DICT)
    workbook.close()

if __name__ == "__main__":
    generate()
