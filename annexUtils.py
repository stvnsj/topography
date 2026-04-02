import re
from datetime import datetime
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl import load_workbook
import numpy as np


"""
openpyxl's iteration.
======================

for merged_range in self.ws.merged_cells.ranges:

iterates over the merged regions of the sheet such as 

[CellRange('A1:B1'), CellRange('D4:D6')]

for merged_range in ws.merged_cells.ranges:
    print(merged_range)              # e.g., A1:B1
    print(merged_range.min_row)      # e.g., 1
    print(merged_range.max_col)      # e.g., 2
    print(merged_range.coord)        # e.g., 'A1:B1'
    print(merged_range.bounds)       # (min_col, min_row, max_col, max_row)


Obtaining the value in the iteration: 

    # Get the top-left cell (anchor cell)
    min_row, min_col = merged_range.min_row, merged_range.min_col
    value = ws.cell(row=min_row, column=min_col).value


"""

INCH_COL = 10
INCH_ROW = 72
OFFSET   = 38
PAGEBREAKS = []


"""
Implement a class that retains the offset, so that
each succesive call starts from the previous range
end.
"""
class HorizontalRange :
    
    def __init__ (self) :
        self.__offset__ = 0
    
    def get_range (self, col1, col2, row1, row2):
        ran = (col1 + self.__offset__ , col2 + self.__offset__, row1, row2)
        self.__offset__ += col2
        return ran
    
    def get_cell (self, col, row) :
        ran = (col + self.__offset__ , row + self.__offset__)
        self.__offset__ += col
        return ran
    
    def reset (self):
        self.__offset__ = 0


def offset_range (col1, col2, pair):
    return (col1 + pair[1], col2 + pair[1])
    


def column_letter_to_number(column: str) -> int:
    column = column.upper()
    column_number = 0
    for char in column:
        column_number = column_number * 26 + (ord(char) - ord('A') + 1)
    return column_number


def column_number_to_letter(number: int) -> str:
    if number < 1:
        raise ValueError("Column number must be 1 or greater.")
    column = ""
    while number > 0:
        number -= 1  # Adjust for 0-based index
        column = chr(number % 26 + ord('A')) + column
        number //= 26
    return column


def cell_of_coor (col : int, row : int) -> str :
    col_letter = column_number_to_letter (col)
    return f'{col_letter}{row}'

def range_of_coor (col1 , row1, col2, row2):
    cell1 = cell_of_coor(col1,row1)
    cell2 = cell_of_coor(col2,row2)
    return f'{cell1}:{cell2}'

def letter_of_int (i) :
    return get_column_letter(i)

def int_of_letter (l) :
    return column_index_from_string(l)

def coordinate_of_ints (i,j) :
    return f'{letter_of_int(i)}:{j}'

class Writer :
    def __init__ (self,workbook,worksheet):
        self.workbook = workbook
        self.worksheet = worksheet
    
    def set_worksheet (self,ws):
        self.worksheet = ws
        
    
    def cell (self, col, row, dat, *dicts):
        cell = cell_of_coor(col,row)
        self.write(cell, dat, *dicts)
    
    def range (self, col1, col2, row1, row2, dat, *dicts):
        ran = range_of_coor(col1,row1,col2,row2)
        self.merge(ran, dat, *dicts)
    
    def merge (self,ran,dat,*dic):
        CELL_FORMAT = {}
        for f in dic :
            CELL_FORMAT = CELL_FORMAT | f
            
        self.worksheet.merge_range(
            ran,
            dat,
            self.workbook.add_format(CELL_FORMAT)
        )
    
    def write (self,cell,dat,*dic):
        
        CELL_FORMAT = {}
        for f in dic :
            CELL_FORMAT = CELL_FORMAT | f
            
        try:
            self.worksheet.write(
                cell,
                dat,
                self.workbook.add_format(CELL_FORMAT)
            )
        except:
            print(cell)
            self.worksheet.write(
                cell,
                -1,
                self.workbook.add_format(CELL_FORMAT)
            )
            

# This is the way to encapsulate functions and
# variables in a module or class. Can be helpful
# to make code more name-consistent.
class Format :
    
    # Functions
    def SIZE(n):
        return {"font_size":n}
    
    def INDENT(n):
        return {'indent':n}
    
    # Variables
    DOT    = {'border':3}
    BOLD   = {"bold":True}
    ITALIC = {"italic":True}
    DEC    = {'num_format': '#,##0.0000000000'}
    INT    = {'num_format': '#,##0'}
    NUM    = {'num_format': '#,##0.000'}
    NUM1   = {'num_format': '#,##0.0'}
    NUM2   = {'num_format': '#,##0.00'}
    TOP    = {"top" :1}
    BTOP   = {"top" :1}
    BBOTTOM = {"bottom" :1}
    BOTTOM = {"bottom":1}
    BLEFT = {"left": 1}
    BRIGHT = {"right": 1}
    BORDER = {"border":1}
    BORDER_THICK = {"border":2}
    LEFT = {"align":"left"}
    RIGHT = {"align":"right"}
    CENTER = {"align":"center"}
    VCENTER = {"valign":"vcenter"}
    COLOR_RED = {'bg_color': '#ff5034'}
    COLOR_GREEN = {'bg_color': '#61df6d'}   




class Formatter :
    
    def __init__(self, ws):
        self.ws = ws
        
    
    def set_cols(self,dic):
        for col in dic:
            size = dic[col]
            self.ws.set_column(col,col,size*INCH_COL) 
            
    
    def set_rows(self,dic):
        for row in dic:
            size = dic[row]
            self.ws.set_row(row,size*INCH_ROW)

def set_column (ws,widths,INCH_COL=10):
    for col , w in enumerate(widths):
        ws.set_column(col,col,w*INCH_COL)

def set_row (ws, heights) :
    for row , h in enumerate(heights) :
        ws.set_row(row,h*INCH_ROW)

def set_row_dict (ws, heights={}) :
    for h in heights:
        ws.set_row(h, heights[h] * INCH_ROW)



def is_g (s):
    if s.startswith("G"):
        return True
    else:
        False


def is_t (s):
    if s.startswith("T"):
        return True
    else:
        return False

def curr_date (opt=0) :
    
    if opt  == 0:
        MES = {
            1 : "ENERO",
            2 : "FEBRERO",
            3 : "MARZO",
            4 : "ABRIL",
            5 : "MAYO",
            6 : "JUNIO",
            7 : "JULIO",
            8 : "AGOSTO",
            9 : "SEPTIEMBRE",
            10 : "OCTUBRE",
            11 : "NOVIEMBRE",
            12 : "DICIEMBRE",
        }
        return f'{MES[datetime.now().month]} {datetime.now().year}'
    
    else :
        MES  = {
            1 : "ENE",
            2 : "FEB",
            3 : "MAR",
            4 : "ABR",
            5 : "MAY",
            6 : "JUN",
            7 : "JUL",
            8 : "AGO",
            9 : "SEP",
            10 : "OCT",
            11 : "NOV",
            12 : "DEC",
        }
        return f'{MES[datetime.now().month]} {datetime.now().year}'


class PRScanner :
    def __init__ (self, ws):
        
        self.ws = ws
        
        # Project data.
        self.PUNTO = None
        self.DM    = None
        self.LADO  = None
        self.DIST  = None
        self.N     = None
        self.E     = None
        self.COTA  = None
        
        # Col letters of the data table.
        self.PROYECTO = None
        self.SECTOR   = None
        self.TRAMO    = None
        self.REALIZADO = None
        self.FECHA     = None
        

        self.HEADER_ROW   = None
        self.MIN_DATA_ROW = None
        self.MAX_DATA_ROW = None
        
        self.__init_row__()
        self.__init_cols__()
        self.__init_project_data__()
    
    
    def __init_row__ (self) :
        self.__init_header_row__()
        self.MIN_DATA_ROW = self.HEADER_ROW + 1
        self.__init_max_data_row__()
        
        
    
    # Initialize MIN_DATA_ROW variable
    def __init_header_row__ (self) :
        for col in self.ws.iter_cols(min_row = 0):
            for cell in col:
                r = cell.row
                if bool(re.match(r"^punto", str(cell.value), re.I)):
                    self.HEADER_ROW = r
                    return
    
    def __init_max_data_row__ (self):
        i = self.MIN_DATA_ROW
        for col in self.ws.iter_cols(min_col=0, min_row = i):
            for cell in col:
                i += 1
            self.MAX_DATA_ROW = i - 1
            return
    
    def __init_cols__ (self):
        for row in self.ws.iter_rows(min_col = 0, min_row = self.HEADER_ROW, max_row = self.HEADER_ROW):
            for cell in row:
                if bool(re.match(r"^punto", str(cell.value), re.I)):
                    self.PUNTO = cell.column
                    continue
                if bool(re.match(r"^dm", str(cell.value), re.I)):
                    self.DM = cell.column
                    continue
                if bool(re.match(r"^lado", str(cell.value), re.I)):
                    self.LADO = cell.column
                    continue
                if bool(re.match(r"^dist", str(cell.value), re.I)):
                    self.DIST = cell.column
                    continue
                if bool(re.match(r"^n", str(cell.value), re.I)):
                    self.N = cell.column
                    continue
                if bool(re.match(r"^e", str(cell.value), re.I)):
                    self.E = cell.column
                    continue
                if bool(re.match(r"^cota", str(cell.value), re.I)):
                    self.COTA = cell.column
                    break

    def __init_project_data__ (self) :
        for col in self.ws.iter_cols(min_col = 0, max_col = 0):
            for cell in col:
                r = cell.row
                c = cell.column
                if bool(re.match(r"^proyecto", str(cell.value), re.I)):
                    self.PROYECTO = self.ws.cell(row=r,column=c+1).value
                    continue
                if bool(re.match(r"^sector", str(cell.value), re.I)):
                    self.SECTOR = self.ws.cell(row=r,column=c+1).value
                    continue
                if bool(re.match(r"^tramo", str(cell.value), re.I)):
                    self.TRAMO = self.ws.cell(row=r,column=c+1).value
                    continue
                if bool(re.match(r"^realizado", str(cell.value), re.I)):
                    self.REALIZADO = self.ws.cell(row=r,column=c+1).value
                    continue
                if bool(re.match(r"^fecha", str(cell.value), re.I)):
                    self.FECHA = self.ws.cell(row=r,column=c+1).value
                    continue
    
    def get_punto(self,row):
        return self.ws.cell(column=self.PUNTO,row=row).value
    def get_dm(self,row):
        return self.ws.cell(column=self.DM,row=row).value
    def get_lado(self,row):
        return self.ws.cell(column=self.LADO,row=row).value
    def get_dist(self,row):
        return self.ws.cell(column=self.DIST,row=row).value
    def get_n(self,row):
        return self.ws.cell(column=self.N,row=row).value
    def get_e(self,row):
        return self.ws.cell(column=self.E,row=row).value
    def get_cota(self,row):
        return self.ws.cell(column=self.COTA,row=row).value  



class Scanner :
    
    def __init__ (self,ws) :
        
        self.ws = ws
        self.row = self.__find_row__()
        
        self.DATUM = None
        self.ZONA  = None
        self.MC    = None

        self.PROYECTO  = None
        self.SECTOR    = None
        self.TRAMO     = None
        self.REALIZADO = None
        
        self.PTL_COLS = []
        
        self.ALTURA_PTL        = []
        self.MERIDIANO_CENTRAL = []
        self.NORTE_FALSO       = []
        self.ESTE_FALSO        = []
        self.FACTOR_ESCALA     = []
        
        self.PRO       = None
        self.EST       = None
        self.GEO_S     = None
        self.GEO_W     = None
        self.UTM_N     = None
        self.UTM_E     = None
        self.GEO_X     = None
        self.GEO_Y     = None
        self.GEO_Z     = None
        self.PTL_N     = []
        self.PTL_E     = []
        self.ELIP      = None
        self.COTA_ORTO = None
        self.COTA_GEO  = None
        self.DM        = None
        self.DIST      = None
        
        self.__find_field__()
        self.__init_ptl__ ()
        self.__init_global__()
    
    def get_pro(self,row):
        return self.ws.cell(column=self.PRO,row=row).value
    
    def get_est(self,row):
        return self.ws.cell(column=self.EST,row=row).value
    
    def get_geo_s(self,row):
        return self.ws.cell(column=self.GEO_S,row=row).value
    
    def get_geo_w (self,row):
        return self.ws.cell(column=self.GEO_W,row=row).value
    
    def get_utm_n (self,row):
        return self.ws.cell(column=self.UTM_N,row=row).value
    
    def get_utm_e  (self,row):
        return self.ws.cell(column=self.UTM_E,row=row).value
    
    def get_geo_x  (self,row):
        return self.ws.cell(column=self.GEO_X,row=row).value
    
    def get_geo_y  (self,row):
        return self.ws.cell(column=self.GEO_Y,row=row).value
    
    def get_geo_z  (self,row):
        return self.ws.cell(column=self.GEO_Z,row=row).value
    
    def get_elip  (self,row):
        return self.ws.cell(column=self.ELIP,row=row).value
    
    def get_cota_orto  (self,row):
        return self.ws.cell(column=self.COTA_ORTO,row=row).value
    
    def get_cota_geo  (self,row):
        return self.ws.cell(column=self.COTA_GEO,row=row).value
    
    def get_dm (self,row):
        return self.ws.cell(column=self.DM, row = row).value
 
    def get_dist (self,row):
        val = self.ws.cell(column=self.DIST, row = row).value
        try:
            return f"{float(val):.2f}"
        except:
            return self.ws.cell(column=self.DIST, row = row).value
 
    def __init_global__ (self) :
        for col in self.ws.iter_cols(min_row = 0, max_row = self.row):
            for cell in col:
                r = cell.row
                c = cell.column
                if bool(re.match(r"^datum", str(cell.value), re.I)):
                    self.DATUM = self.ws.cell(row=r,column=c+1).value
                    continue
                if bool(re.match(r"^zona", str(cell.value), re.I)):
                    self.ZONA = self.ws.cell(row=r,column=c+1).value
                    continue
                if bool(re.match(r"^mc", str(cell.value), re.I)):
                    self.MC = self.ws.cell(row=r, column=c+1).value
                    continue
                if bool(re.match(r"^proyecto", str(cell.value), re.I)):
                    self.PROYECTO = self.ws.cell(row=r, column=c+1).value
                    continue
                if bool(re.match(r"^sector", str(cell.value), re.I)):
                    self.SECTOR = self.ws.cell(row=r, column=c+1).value
                    continue
                if bool(re.match(r"^tramo", str(cell.value), re.I)):
                    self.TRAMO = self.ws.cell(row=r, column=c+1).value
                    continue
                if bool(re.match(r"^realizado", str(cell.value), re.I)):
                    self.REALIZADO = self.ws.cell(row=r, column=c+1).value
                    continue
    
    
    def get_t_rows (self) :
        t_rows = []
        for col in self.ws.iter_cols(min_col=self.EST, max_col=self.EST, min_row=self.row):
                for cell in col:
                    if re.match(r"^t", str(cell.value), re.I):
                        t_rows.append(cell.row)
        return t_rows
    
    def get_g_rows (self) :
        g_rows = []
        for col in self.ws.iter_cols(min_col=self.EST, max_col=self.EST, min_row=self.row):
                for cell in col:
                    if re.match(r"^g", str(cell.value), re.I):
                        g_rows.append(cell.row)
        return g_rows
        
    
    def get_poligonal_num (self, row):
        target_cell = self.ws.cell(row = row, column = self.PRO)
        for merged_range in self.ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                cell = self.ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                match = re.search(r'\d+', str(cell.value))
                if match:
                    return int(match.group())
                else:
                    return -1
    
    def __init_ptl__ (self):
        for i in self.PTL_COLS:
            for col in self.ws.iter_cols(min_col=i,max_col=i,min_row=1,max_row=self.row):
                for cell in col:
                    if bool(re.match(r"altura", str(cell.value), re.I)):
                        self.ALTURA_PTL.append(self.ws.cell(column = cell.column + 1 , row = cell.row).value)
                        continue
                    if bool(re.match(r"meridiano", str(cell.value), re.I)):
                        self.MERIDIANO_CENTRAL.append(self.ws.cell(column = cell.column + 1 , row = cell.row).value)
                        continue
                    if bool(re.match(r"norte", str(cell.value), re.I)):
                        self.NORTE_FALSO.append(self.ws.cell(column = cell.column + 1 , row = cell.row).value)
                        continue
                    if bool(re.match(r"este", str(cell.value), re.I)):
                        self.ESTE_FALSO.append(self.ws.cell(column = cell.column + 1 , row = cell.row).value)
                        continue
                    if bool(re.match(r"factor", str(cell.value), re.I)):
                        self.FACTOR_ESCALA.append(self.ws.cell(column = cell.column + 1 , row = cell.row).value)
                        continue
    
 
    def __find_row__ (self) :
        for col in self.ws.iter_cols(min_col=1):
            for cell in col:
                if cell.value == "MASTER":
                    return cell.row
        raise Exception("No row starting with \'MASTER\' cell")
    
    def __find_field__ (self) :
        for row in self.ws.iter_rows(min_row=self.row,max_row=self.row):
            for cell in row:
                if cell.value == "PRO":
                    self.PRO = cell.column
                    continue
                if cell.value == "EST":
                    self.EST = cell.column
                    continue
                if cell.value == "GEO-S":
                    self.GEO_S = cell.column
                    continue
                if cell.value == "GEO-W":
                    self.GEO_W = cell.column
                    continue
                if cell.value == "UTM-N":
                    self.UTM_N = cell.column
                    continue
                if cell.value == "UTM-E":
                    self.UTM_E = cell.column
                    continue
                if cell.value == "GEO-X":
                    self.GEO_X = cell.column
                    continue
                if cell.value == "GEO-Y":
                    self.GEO_Y = cell.column
                    continue
                if cell.value == "GEO-Z":
                    self.GEO_Z = cell.column
                    continue
                if  bool(re.match(r"^PTL\d*-N", str(cell.value))):
                    self.PTL_COLS.append(cell.column)
                    self.PTL_N.append(cell.column)
                    continue
                if  bool(re.match(r"^PTL\d*-E", str(cell.value))):
                    self.PTL_E.append(cell.column)
                    continue
                if cell.value == "ELIP":
                    self.ELIP = cell.column
                    continue
                if cell.value == "COTA-ORTO":
                    self.COTA_ORTO = cell.column
                    continue
                if cell.value == "COTA-GEO":
                    self.COTA_GEO =  cell.column
                    continue
                if cell.value == "DM":
                    self.DM = cell.column
                    continue
                if cell.value == "DIST":
                    self.DIST = cell.column
                    break



class Scanner10 :
    
    def __init__ (self, ws) :
        self.ws = ws
        self.PROYECTO = None
        self.SECTOR   = None
        self.TRAMO    = None
        self.REALIZADO = None
    
    def __init_cols__ (self) :
        pass
    
    
    
    


if __name__ == "__main__":
    wb = load_workbook("/home/jstvns/axis/axis/anexos/anteproyecto/anexo0.xlsx")
    ws = wb.active
    scanner = PRScanner(ws)
    print(scanner.get_cota(12))
    
