

import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
import sys
import reader as rd
import re
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
import annexUtils
from openpyxl import load_workbook

import os
import glob
import annexImg

OFFSET   = 30
PAGEBREAKS = []

def generate (input_file='anexos/anteproyecto/anexo1.xlsx',output_file="test5.xlsx", src_dir="", src_dir2="img_geo",crop_zoom=1.5) :
    
    print(f'\n\nGeneración de {output_file} en curso ...')
    
    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet("Fichas PR")
    writer = Writer(workbook,worksheet)
    wb = load_workbook(input_file)
    ws = wb.active
    scanner = annexUtils.PRScanner(ws)
    dst_dir2 = None
    
    COL_WIDTHS = [
        0.10, #A 
        0.30, #B
        0.70, #C
        0.25, #D
        0.39, #E
        0.18, #F
        0.19, #G
        0.19, #H
        0.25, #I
        0.22, #J
        0.30, #K
        0.25, #L
        0.15, #M
        0.42, #N
        0.15, #O
        0.20, #P
        0.20, #Q
        0.33, #R
        0.25, #S
        0.15, #T
        0.25, #U
        0.15, #V
        0.33, #W
        0.28, #X
        0.46, #Y
        0.25, #Z
        0.10, #AA
    ]
    
    ROW_DICT = {
        0 :0.1,  
        6 :0.1,  
        7 :0.18, 
        8 :0.18, 
        9 :0.19, 
        10:0.1,
        12:0.12,
        13:0.12
        
    }
    
    #worksheet.autofit()
    annexUtils.set_column(worksheet,COL_WIDTHS)
    
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    
    # FIXED CONTENT
    writer.merge(f"B2:F6","",Format.BORDER)
    writer.merge(f"G2:Z3","MONOGRAFÍAS DE PR",
                 {"top":1, "right":1,"font_size":12,"bold":True,"align":"center","valign":"vcenter"})
    
    writer.merge(f"G4:Z6","FORMULARIO N° 2.903.3.I   FIGURA 3",
                 {"bottom":1,"right":1, "font_size":12,"bold":True,"align":"center","valign":"vcenter"})
    
    writer.write(f"AA2","",Format.BLEFT)
    
    writer.merge(f"B7:Z7","",{"bottom":1})
    writer.merge(f"A8:A12","",{"right":1})
    writer.merge(f"AA8:AA12","",{"left":1})
    writer.merge(f"B13:Z13", "",{"top":1})
    
    writer.merge(f"B8:C8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B9:C9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B10:C10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B12:C12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"U12:Z12",f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
    
    writer.merge(f'D8:Y8', scanner.PROYECTO, Format.LEFT,Format.SIZE(10))
    writer.merge(f'D9:Y9',scanner.SECTOR,Format.SIZE(10))
    writer.merge(f'D10:Y10', scanner.TRAMO,Format.SIZE(10))
    writer.merge(f'D12:R12',scanner.REALIZADO,Format.SIZE(10))
    
    min_data_row = scanner.MIN_DATA_ROW
    max_data_row = scanner.MAX_DATA_ROW
  
    i = 0
    
    src_dir = None
        
    if src_dir2:
        dst_dir2 = annexImg.annex9_process_geo(src_dir2,crop_zoom=crop_zoom)
    
    
    # LOOP THE FOLLOWING CELLS 
    for r in range(min_data_row, max_data_row+1) :
        pr_punto = scanner.get_punto(r)
        pr_dm    = scanner.get_dm(r)
        pr_lado  = scanner.get_lado(r)
        pr_dist  = scanner.get_dist(r)
        pr_n     = scanner.get_n(r)
        pr_e     = scanner.get_e(r)
        pr_cota  = scanner.get_cota(r)
        
        writer.write(f"B{15 + i * OFFSET}","Identificación del Punto",Format.BOLD, Format.ITALIC, Format.SIZE(10))
        writer.write(f"F{15 + i * OFFSET}", "PR:", Format.SIZE(10))
        writer.write(f"K{15 + i * OFFSET}", "Dm. Ref.:",Format.SIZE(10))
        writer.write(f"R{15 + i * OFFSET}", "Lado:")
        
        
        writer.write(f'F{16 + i * OFFSET}', "Cota:",Format.SIZE(10))
        writer.merge(f"H{16 + i * OFFSET}:J{16 + i * OFFSET}", pr_cota, Format.NUM)
        writer.write(f"K{16 + i * OFFSET}", "m")
        
        writer.write(f'M{16 + i * OFFSET}',"Coordenadas de Navegación UTM:",Format.SIZE(10))
        
        writer.write(f'V{16 + i * OFFSET}', "N:")
        writer.write(f'V{17 + i * OFFSET}', "E:")
        
        writer.merge(f'W{16 + i * OFFSET}:Z{16 + i * OFFSET}', pr_n, Format.NUM)
        writer.merge(f'W{17 + i * OFFSET}:Z{17 + i * OFFSET}', pr_e, Format.NUM)
        
        
        
        
        writer.merge(
            f"H{15 + i * OFFSET}:I{15 + i * OFFSET}",
            pr_punto,
            Format.SIZE(11), Format.BOTTOM, Format.CENTER
        )
        
        
        writer.merge(f"M{15 + i * OFFSET}:P{15 + i * OFFSET}",pr_dm,Format.CENTER,Format.SIZE(10),Format.BOTTOM,Format.NUM2)
        writer.merge(f"T{15 + i * OFFSET}:U{15 + i * OFFSET}", pr_lado, Format.CENTER, Format.BBOTTOM)
        writer.write(f"W{15 + i * OFFSET}","FECHA:",Format.SIZE(10))
        writer.merge(f"Y{15 + i * OFFSET}:Z{15 + i * OFFSET}", annexUtils.curr_date(1),Format.BOTTOM,Format.CENTER,Format.SIZE(10))
        
        
        
        
        
        if src_dir2 :

            print("DIRECTORIO")
            print(dst_dir2)

            cleaned_point = pr_punto.replace("-", "")


            img_path_a = os.path.join(dst_dir2, f'{cleaned_point}_a.*')
            img_path_p = os.path.join(dst_dir2, f'{cleaned_point}_p.*')
            img_path_g = os.path.join(dst_dir2, f'{cleaned_point}_g.*')
            
            match_a = glob.glob(img_path_a)
            match_p = glob.glob(img_path_p)
            match_g = glob.glob(img_path_g)
            
            if match_a:
                worksheet.insert_image(f'B{33 + i * OFFSET}', match_a[0] , {'object_position': 1})
            else :
                writer.merge(
                    f"B{33 + i * OFFSET}:F{40 + i * OFFSET}","Fotografía\nDetalle",
                    Format.BORDER,
                    Format.CENTER,Format.VCENTER)
            
            if match_p:
                worksheet.insert_image(f'B{19 + i * OFFSET}', match_p[0],  {'object_position': 1})
                
            else:
                writer.merge(
                    f"B{19 + i * OFFSET}:L{31 + i * OFFSET}","Fotografía\nPanorámica",
                    Format.BORDER,
                    Format.CENTER,Format.VCENTER)
     
            if match_g:
                worksheet.insert_image(f'N{19 + i * OFFSET}', match_g[0] , {'object_position': 1})
            else:
                writer.merge(
                    f"N{19 + i * OFFSET}:Z{31 + i * OFFSET}","Vista\nAérea",
                    Format.BORDER,
                    Format.CENTER,Format.VCENTER)
        
        
        writer.merge(f"H{41-8 + i * OFFSET}:Z{41-8 + i * OFFSET}","Descripción",Format.BOTTOM,Format.LEFT,Format.SIZE(10))
        
        writer.write(f'I{43-8 + i * OFFSET}', "Materialidad:",Format.SIZE(9))
        writer.write(f'I{44-8 + i * OFFSET}', "Dimensiones:",Format.SIZE(9))
        writer.merge(f'I{45-8 + i * OFFSET}:M{45-8 + i * OFFSET}', "Distancia a la Ruta:",Format.SIZE(9))
        writer.merge(f'N{45-8 + i * OFFSET}:O{45-8 + i * OFFSET}', pr_dist, Format.SIZE(9), Format.NUM2)
        writer.write(f'P{45-8 + i * OFFSET}', "m", Format.SIZE(9), Format.LEFT)
        
        writer.write(f'L{43-8 + i * OFFSET}', "MONOLITO DE HORMIGÓN, PINTADO DE AMARILLO", Format.SIZE(9))
        writer.write(f'L{44-8 + i * OFFSET}', "30,0 X 30,0 X 50,0 cm.", Format.SIZE(9))
        writer.write(f'M{47-8 + i * OFFSET}', "SENTIDO DE AVANCE", Format.SIZE(9))
        
        writer.merge(f"G{42-8 + i * OFFSET}:G{48-8 + i * OFFSET}","",Format.BRIGHT)
        writer.merge(f"AA{42-8 + i * OFFSET}:AA{48-8 + i * OFFSET}","",Format.BLEFT)
        writer.merge(f"H{49-8 + i * OFFSET}:Z{49-8 + i * OFFSET}","",Format.TOP)
        writer.merge(f"A{50-8 + i * OFFSET}:AA{50-8 + i * OFFSET}","",{})
        
        ROW_DICT.update({15 : 0.33})
        
        
        ROW_DICT.update({key:0.16 for key in range(19 + i*OFFSET , 31 + i*OFFSET)})  # - 8
        ROW_DICT.update({key:0.175 for key in range(33 + i*OFFSET , 40 + i*OFFSET)})  
        
        PAGEBREAKS.append(50-8 + i * OFFSET)
        
        i += 1
    
    worksheet.set_h_pagebreaks(PAGEBREAKS)
    annexUtils.set_row_dict(worksheet,ROW_DICT)
    
    workbook.close()
    
    print(f'\nGeneración de {output_file} completada\n')

if __name__ == "__main__":
    generate(input_file='/home/jstvns/axis/axis/anexos/anteproyecto/anexo0.xlsx', output_file="../test9.xlsx", src_dir = '../img', src_dir2= '../geo')
