import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
import sys
import reader as rd
import re
import annexUtils
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
from openpyxl   import load_workbook

class Polygone :
    def __init__ (self, coordinate, min_row, max_row, value):
        self.coordinate = coordinate
        self.min_row = min_row
        self.max_row = max_row
        self.value = value
        
    def __lt__ (self,p) :
        return self.min_row < p.min_row
    
    def __le__ (self,p) :
        return self.min_row <= p.min_row
    
    def __str__ (self)  :
        return f'coor: {self.coordinate}\nmin_row: {self.min_row}\nmax_row: {self.max_row}\nval: {self.value}'


def generate (input_file = 'anexos/anteproyecto/annex1.xlsx', output_file='test8.xlsx') :
    
    workbook = xlsxwriter.Workbook(output_file)
    wb = load_workbook(input_file)
    ws = wb.active
    scanner = annexUtils.Scanner(ws)
    column = 'B'
    
    polygones = [] 
    
    for merged_range in ws.merged_cells.ranges:
        
        if merged_range.min_col == openpyxl.utils.column_index_from_string(column):
            
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            if top_left_cell.value.startswith("POLIGONAL"):
                polygone = Polygone(
                    top_left_cell.coordinate,
                    merged_range.min_row,
                    merged_range.max_row,
                    top_left_cell.value
                )
                polygones.append(polygone)
    
    polygones.sort()
    
    # for p in polygones:
    #     print("")
    #     print(p)
    #     print("")
    
    
    COL_WIDTHS = [
        0.09, #A
        0.67, #B
        0.13, #C
        0.20, #D
        0.15, #E
        0.55, #F
        0.20, #G
        0.55, #H
        0.14, #I
        0.30, #J
        0.14, #K
        0.25, #L
        0.30, #M
        0.20, #N
        0.54, #O
        0.20, #P
        0.20, #Q
        0.20, #R
        0.20, #S
        0.20, #T
        0.20, #U
        0.70, #V
        0.09, #W
    ]


    
    ROW_HEIGHTS = [
        0.12, #1
        0.21, #2  
        0.21, #3
        0.21, #4
        0.21, #5
        0.21, #6
        0.09, #7
        0.17, 0.17, 0.17, # 8,9,10
        0.07, #11
    ]
    
    for p in polygones:
        
        worksheet = workbook.add_worksheet(p.value)
        
        annexUtils.set_column(worksheet,COL_WIDTHS)
        annexUtils.set_row(worksheet,ROW_HEIGHTS)
        
        worksheet.hide_gridlines(2)
        worksheet.set_portrait()
        worksheet.set_page_view(2)
        worksheet.set_paper(9)
        worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)
        
        writer = Writer(workbook,worksheet)
        
        # FIXED CONTENT
        writer.merge(f"B2:G6","",Format.BORDER)

        writer.merge(
            f"H2:V3","COORDENADAS DE VÉRTICES DEL STC",
            Format.BTOP,Format.BRIGHT,Format.SIZE(12),Format.BOLD,
            Format.CENTER, Format.VCENTER
        )
        
        writer.merge(
            f"H4:V6",
            "FORMULARIO N° 2.303.104.B",
            Format.BBOTTOM,Format.BRIGHT,Format.SIZE(12),
            Format.BOLD,Format.CENTER,Format.VCENTER
        )
        
        
        writer.write("W2","",Format.BLEFT)
        writer.merge("B7:V7","",{"bottom":1})
        writer.merge("A8:A12","",{"right":1})
        writer.merge("W8:W12","",{"left":1})
        writer.merge("B13:V13", "",{"top":1})






        
        
        writer.write("B8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
        writer.write("B9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
        writer.write("B10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
        writer.write("B12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
        writer.merge("Q12:V12",f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
        
        writer.merge(f'D8:V8', scanner.PROYECTO,Format.SIZE(10))
        writer.merge(f'D9:V9',scanner.SECTOR,Format.SIZE(10))
        writer.merge(f'D10:V10', scanner.TRAMO,Format.SIZE(10))
        writer.merge(f'D12:O12',scanner.REALIZADO,Format.SIZE(10))
        
        writer.merge(
            'B15:B18','Vértice',Format.SIZE(9),
            Format.BOLD,Format.CENTER,Format.VCENTER, Format.BORDER
        )
        
        ROW = p.min_row + 1
        writer.merge(
            'B14:Q14', f'POLIGONAL Nº: {scanner.get_poligonal_num(ROW)}      Tipo: Principal',
            Format.SIZE(9),Format.BOLD,Format.BORDER,Format.CENTER,Format.VCENTER,
        )
        
        for idx, C in enumerate(scanner.PTL_N):
            LTM_N = ws.cell(column=scanner.PTL_N[idx],row=ROW).value
            try:
                float(LTM_N)
                if len(scanner.PTL_N) == 1:
                    writer.merge( 'D15:I15','PTL',Format.SIZE(9),Format.BOLD,Format.CENTER)
                else:
                    writer.merge( 'D15:I15',f'PTL-{idx+1}',Format.SIZE(9),Format.BOLD,Format.CENTER)
                writer.merge( 'F16:I16',scanner.MERIDIANO_CENTRAL[idx],Format.SIZE(9), Format.BOLD, Format.LEFT)
                writer.merge( 'F17:I17',scanner.FACTOR_ESCALA[idx],Format.SIZE(9), Format.BOLD, Format.LEFT)
                break
            except:
                continue
        
        
        writer.merge( 'D16:E16','MCL:',Format.SIZE(9), Format.BOLD, Format.LEFT)
        writer.merge( 'D17:E17','Kptl:',Format.SIZE(9), Format.BOLD, Format.LEFT)
        
        writer.merge( 'J15:J17','', Format.BRIGHT)
        writer.merge( 'C18:J18','', Format.BBOTTOM,Format.BRIGHT)
        
        writer.merge( 'L16:P16','UTM',Format.SIZE(9), Format.BOLD, Format.CENTER)
        writer.merge( 'L17:P17',f'Huso: {scanner.ZONA}',Format.SIZE(9), Format.BOLD, Format.CENTER)
        writer.merge( 'K18:Q18', '', Format.BBOTTOM,Format.BRIGHT)
        writer.merge( 'Q15:Q17', '', Format.BRIGHT)
        
        curr_row = 19
        offset   = 4
        
        for i in range(p.min_row+1, p.max_row):
            
            POINT = scanner.get_est(i)
            NL    = None
            EL    = None
            N     = scanner.get_utm_n(i)
            E     = scanner.get_utm_e(i)
            COTA  = scanner.get_cota_geo(i)
            H     = scanner.get_cota_orto(i)
            
            writer.merge(
                f'B{curr_row}:B{curr_row+3}',POINT,Format.SIZE(9),
                Format.BOLD,Format.CENTER,Format.VCENTER, Format.BORDER
            )
            
            writer.merge(f'D{curr_row}:E{curr_row}', "NL:",Format.SIZE(10))
            writer.merge(f'D{curr_row+1}:E{curr_row+1}', "EL:",Format.SIZE(10))
            writer.merge(f'D{curr_row+2}:F{curr_row+2}', "Cota (nivelada):",Format.SIZE(9))
            
            for idx, C in enumerate(scanner.PTL_N):
                LTM_N = ws.cell(column=scanner.PTL_N[idx],row=i).value
                try:
                    float(LTM_N)
                    NL = ws.cell(column=scanner.PTL_N[idx],row=i).value
                    EL = ws.cell(column=scanner.PTL_E[idx],row=i).value
                    break
                except:
                    continue
            
            writer.merge(f'F{curr_row}:H{curr_row}',NL,Format.RIGHT,Format.SIZE(10),Format.NUM)
            writer.merge(f'F{curr_row+1}:H{curr_row+1}',EL,Format.RIGHT,Format.SIZE(10), Format.NUM)
            
            
            writer.merge(f'G{curr_row+2}:H{curr_row+2}',COTA,Format.RIGHT,Format.SIZE(10), Format.NUM)
            
            writer.write(f'I{curr_row}','m')
            writer.write(f'I{curr_row+1}','m')
            writer.write(f'I{curr_row+2}','m')
            
            writer.merge(f'C{curr_row+3}:J{curr_row+3}','',Format.BOTTOM,Format.BRIGHT)
            writer.merge(f'J{curr_row}:J{curr_row+2}','',Format.BRIGHT)
            
            
            
            writer.write(f'L{curr_row}',"N:",Format.LEFT,Format.SIZE(10))
            writer.write(f'L{curr_row+1}',"E:",Format.LEFT,Format.SIZE(10))
            writer.merge(f'L{curr_row+2}:M{curr_row+2}',"H(model):",Format.LEFT,Format.SIZE(9))
            
            writer.merge(f'M{curr_row}:O{curr_row}', N, Format.RIGHT,Format.SIZE(10), Format.NUM)
            writer.merge(f'M{curr_row+1}:O{curr_row+1}', E, Format.RIGHT,Format.SIZE(10), Format.NUM)
            writer.merge(f'N{curr_row+2}:O{curr_row+2}', H, Format.RIGHT,Format.SIZE(10),Format.NUM)
            
            writer.merge(f'Q{curr_row}:Q{curr_row+2}', '', Format.BRIGHT)
            writer.merge(f'K{curr_row+3}:Q{curr_row+3}', '',Format.BBOTTOM,Format.BRIGHT)
            
            
            writer.write(f'P{curr_row}'  , 'm', Format.RIGHT)
            writer.write(f'P{curr_row+1}', 'm', Format.RIGHT)
            writer.write(f'P{curr_row+2}', 'm', Format.RIGHT)
            
            
            curr_row += offset
            
        
        writer.merge(f'B{curr_row}:Q{curr_row}','',Format.BTOP)
        
        
        
        
    
    workbook.close()

if __name__ == "__main__":
    generate()
