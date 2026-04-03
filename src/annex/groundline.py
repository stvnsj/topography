import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
from utils import str_to_flt
import sys
import reader as rd
import re
import annexUtils
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
from openpyxl   import load_workbook
from control.mop import MOPControl
from control.level import LevelControl
import level

ROW_DICT = {}

def generate (
        f1_mop = "",
        f2_mop = "",
        f1_long = '',
        f2_long = '',
        output_file = "",
        dm_interval = "*",
        total_length = 1000,
        seed = 42
) :
    
    if f1_mop == "":
        print("Error: Archivo MOP de proyecto no cargado")
        return
    if f2_mop == "":
        print("Error: Archivo MOP de control no cargado")
        return
    if f1_long == "":
        print("Error: Archivo longitudinal de proyecto no cargado")
        return
    if f2_long == "":
        print("Error: Archivo longitudinal de control no cargado")
        return
    
    
    
    workbook   = xlsxwriter.Workbook(output_file)
    worksheet  = workbook.add_worksheet("autocontrol")
    
    level_control = LevelControl(f1_long,f2_long)
    random_mop = MOPControl(f1_mop,f2_mop).select_random_points(seed=seed)
    
    
    worksheet.hide_gridlines(2)  # 2 hides both the printed and visible gridlines
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(5)
    
    writer = Writer(workbook,worksheet)
 
    
    writer.range(2,19,2,6,"",Format.BORDER)
    writer.range(20,49,2,4,"VERIFICACIÓN DE LÍNEA DE TIERRA",
                 Format.SIZE(12),Format.BOLD, Format.BTOP,Format.BRIGHT, Format.CENTER, Format.VCENTER)
    writer.range(20,49,5,6,"FORMULARIO N° 2.309.306.A",
                 Format.SIZE(12),Format.BOLD, Format.BBOTTOM,Format.BRIGHT, Format.BOTTOM, Format.CENTER)
 
    
    writer.range(1,1,8,11,"",Format.BRIGHT)
    writer.range(50,50,8,11,"",Format.BLEFT)
    writer.range(2,49,7,7,"",Format.BBOTTOM)
    writer.range(2,49,12,12,"",Format.BTOP)
    
    
    writer.cell (2,8,"PROYECTO:",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.cell (2,9,"SECTOR:", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.cell (2,10,"TRAMO:", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.cell (2,11,"REALIZADO:",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.range(36,49,11,11,f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
 
    
    writer.cell(10,8, ": *")
    writer.cell(10,9, ": *")
    writer.cell(10,10,": *")
    writer.cell(10,11,": AUTOCONTROL TOPOGRÁFICO")
 
    
    writer.range(2,6,14,14,"Tramo N°:",Format.SIZE(9))
    writer.range(7,11,14,14,dm_interval,Format.SIZE(9),Format.CENTER, Format.BBOTTOM) #PROGRAM
    writer.range(13,17,14,14,"Dm. Inicio:" ,Format.SIZE(9))
    writer.range(18,23,14,14,random_mop.min_ctrl_dm ,Format.SIZE(9),Format.CENTER, Format.BBOTTOM) #PROGRAM
    writer.range(25,28,14,14,"Dm. Fin:" ,Format.SIZE(9))
    writer.range(29,33,14,14,random_mop.max_ctrl_dm,Format.SIZE(9),Format.CENTER, Format.BBOTTOM) #PROGRAM
    writer.range(35,42,14,14,"Longitud del Tramo:",Format.SIZE(9))
    writer.range(43,49,14,14,random_mop.ctrl_length(),Format.SIZE(9),Format.CENTER, Format.BORDER) #PROGRAM
    
    
    writer.range(40,45,15,15,"% muestral:", Format.SIZE(9))
    writer.range(46,49,15,15,np.round(100*random_mop.ctrl_length()/total_length,1),Format.SIZE(9),Format.BORDER,Format.NUM1,Format.CENTER)
    
    writer.range(2,40,16,16,"Características del Tramo:",Format.SIZE(9))
    writer.range(2,49,17,19,"",Format.SIZE(9),Format.BORDER)
    
    writer.range(2,15,21,21,"N° de Perfiles Levantados en el Tramo:",Format.SIZE(9))
    writer.range(16,18,21,21,random_mop.section_number(),Format.BORDER,Format.SIZE(9),Format.BORDER,Format.CENTER)
    writer.range(21,30,21,21,"N° de Perfiles Requeridos:",Format.SIZE(9))
    writer.range(31,33,21,21,random_mop.required_sections_per_km(),Format.BORDER,Format.SIZE(9),Format.CENTER)
    writer.range(35,38,21,21,"Cumple:",Format.SIZE(9))
    writer.range(39,42,21,21,
                 "SI" if random_mop.sufficient_points() else 'NO',Format.BORDER,Format.SIZE(9),Format.CENTER)
    
    
    index = 23
    writer.range(2,17,index,index, "Línea de Tierra Longitudinal", Format.BOLD, Format.SIZE(10))
    index+=1
    writer.range(10,19,index,index,"Cota Estacado",Format.BBOTTOM,Format.SIZE(9),Format.CENTER)
    writer.range(25,33,index,index,"Tolerancia (m):",Format.SIZE(9),Format.RIGHT)
    writer.range(34,36,index,index,"0.010",Format.BORDER,Format.SIZE(9))
    index+=1
    
    writer.range(2,4,index,index, "N°",Format.CENTER,Format.SIZE(9))
    writer.range(5,9,index,index, "Dm.", Format.SIZE(9),Format.CENTER)
    writer.range(10,15,index,index,"Estudio",Format.SIZE(9),Format.CENTER)
    writer.range(16,19,index,index,"Control",Format.SIZE(9),Format.CENTER)
    writer.range(20,23,index,index,"Dif. (m)",Format.SIZE(9),Format.CENTER)
    writer.range(25,29,index,index,"Cumple (S/N)",Format.SIZE(9),Format.CENTER)
    writer.range(31,49,index,index,"Observaciones",Format.SIZE(9),Format.CENTER)
    
 
    
    
    index += 1
    level_point_number = 1
    for p in level_control.point_list:
        writer.range(2,4,index,index,level_point_number,Format.SIZE(9),Format.BORDER,Format.CENTER)
        writer.range(5,9,index,index,str_to_flt(p.dm),Format.SIZE(9),Format.BORDER,Format.NUM,Format.CENTER)
        writer.range(10,15,index,index,str_to_flt(p.proj_height),Format.SIZE(9),Format.BORDER, Format.NUM,Format.CENTER)
        writer.range(16,19,index,index,str_to_flt(p.ctrl_height),Format.SIZE(9),Format.BORDER, Format.NUM,Format.CENTER)
        writer.range(20,23,index,index,str_to_flt(p.delta),Format.SIZE(9),Format.BORDER, Format.NUM,Format.CENTER)
        writer.range(25,29,index,index,p.good,Format.SIZE(9),Format.BORDER,Format.CENTER)
        writer.range(31,49,index,index,"",Format.SIZE(9),Format.BORDER,Format.CENTER)
        level_point_number +=1
        index += 1
    
    
    index+=1
    
    writer.range(2,49,index,index,"Observaciones generales:",Format.SIZE(9))
    
    index+=1
    
    writer.range(2,49,index,index+1,"",Format.SIZE(9),Format.BORDER)
    
    index += 4
    
    writer.range(2,17,index,index, "Línea de Tierra Transversal", Format.BOLD, Format.SIZE(10))
    
    index+=1
    writer.range(2,14,index,index, "N° puntos contrastados:", Format.SIZE(10))
    writer.range(15,18,index,index, random_mop.ctrl_number, Format.CENTER, Format.BBOTTOM, Format.SIZE(10)) #PROGRAM
    writer.range(23, 34, index, index, "Puntos en Tolerancia:", Format.SIZE(10))
    writer.range(35, 39, index, index, random_mop.good_number, Format.CENTER, Format.BBOTTOM, Format.SIZE(10)) #PROGRAM
    writer.range(42,44, index, index, "%", Format.RIGHT, Format.SIZE(10))
    writer.range(45,48, index, index, random_mop.good_percent, Format.CENTER, Format.BBOTTOM, Format.SIZE(10)) #PROGRAM
    
    index+=3
    writer.range(2,5,index,index,  "Perfil N°",Format.SIZE(9),Format.INDENT(1))
    writer.range(6,9,index,index,  "Dm.",Format.SIZE(9),Format.INDENT(1))
    writer.range(10,13,index,index,"Lado",Format.SIZE(9),Format.INDENT(1))
    
    writer.range(14,18,index-1,index-1,"Dist.",Format.SIZE(9), Format.INDENT(1))
    writer.range(14,18,index,index,"al Eje",Format.SIZE(9), Format.INDENT(1))
    
    writer.range(19,23,index-1,index-1,"Cota",Format.SIZE(9), Format.INDENT(1))
    writer.range(19,23,index,index, "Control",Format.SIZE(9), Format.INDENT(1))
    
    writer.range(24,28,index-1,index-1,"Cota",Format.SIZE(9), Format.INDENT(1))
    writer.range(24,28,index,index,"Estudio",Format.SIZE(9), Format.INDENT(1))
    
    writer.range(29,33,index-1,index-1,"Tipo",Format.SIZE(9), Format.INDENT(1))
    writer.range(29,33,index,index,"Sup. (1-4)",Format.SIZE(9), Format.INDENT(1))
    
    writer.range(34,38,index,index, "Tol. (m)",    Format.SIZE(9), Format.INDENT(1))
    writer.range(39,42,index,index, "Dif. (m)",    Format.SIZE(9), Format.INDENT(1))
    writer.range(44,49,index,index, "Cumple (S/N)",Format.SIZE(9), Format.INDENT(1))
    
    index += 1
    section_number = 1
    
    for section in random_mop.section_list :
        
        initial_index = index
        
        pos_points = section.pos_points
        neg_points = section.neg_points
        
        writer.range(2,5,index,index,   section_number, Format.SIZE(9), Format.BORDER)
        writer.range(6,9,index,index,   str_to_flt (section.dm), Format.SIZE(9), Format.BORDER, Format.NUM)
        
        writer.range(10,13,
                     index ,
                     index + len(neg_points)-1,
                     "IZQ", Format.SIZE(9),
                     Format.BORDER, Format.CENTER,
                     Format.VCENTER) if neg_points else None
        
        writer.range(10,13,
                     index + len(neg_points),
                     index + len(neg_points) + len(pos_points) - 1,
                     "DER", Format.SIZE(9),
                     Format.BORDER,
                     Format.CENTER,
                     Format.VCENTER) if pos_points else None

        
        for point in neg_points:
            writer.range(14,18,index,index,str_to_flt (point.distance), Format.SIZE(9), Format.BORDER, Format.NUM, Format.CENTER)
            writer.range(19,23,index,index,str_to_flt (point.ctrl_height), Format.SIZE(9), Format.BORDER, Format.NUM, Format.CENTER)
            writer.range(24,28,index,index,str_to_flt (point.proj_height), Format.SIZE(9), Format.BORDER, Format.NUM, Format.CENTER)
            writer.range(29,33,index,index,str_to_flt (point.type), Format.SIZE(9), Format.BORDER, Format.NUM, Format.CENTER)
            writer.range(34,38,index,index,str_to_flt (point.tol), Format.SIZE(9), Format.BORDER, Format.NUM, Format.CENTER)
            writer.range(39,42,index,index,str_to_flt (point.dif), Format.SIZE(9), Format.BORDER, Format.NUM, Format.CENTER)
            writer.range(44,49,index,index, "SI" if point.good == "True" else "NO", Format.SIZE(9), Format.BORDER, Format.CENTER)
            writer.cell (50,index,"",Format.BLEFT)
            index += 1
        
        for point in pos_points:
            writer.range(14,18,index,index,str_to_flt( point.distance), Format.SIZE(9), Format.BORDER, Format.NUM, Format.CENTER)
            writer.range(19,23,index,index,str_to_flt( point.ctrl_height), Format.SIZE(9), Format.BORDER, Format.NUM, Format.CENTER)
            writer.range(24,28,index,index,str_to_flt( point.proj_height), Format.SIZE(9), Format.BORDER, Format.NUM, Format.CENTER)
            writer.range(29,33,index,index,str_to_flt( point.type), Format.SIZE(9), Format.BORDER, Format.NUM, Format.CENTER)
            writer.range(34,38,index,index,str_to_flt( point.tol), Format.SIZE(9), Format.BORDER, Format.NUM, Format.CENTER)
            writer.range(39,42,index,index,str_to_flt( point.dif), Format.SIZE(9), Format.BORDER, Format.NUM, Format.CENTER)
            writer.range(44,49,index,index, "SI" if point.good == "True" else "NO", Format.SIZE(9), Format.BORDER,Format.CENTER)
            writer.cell (50,index,"",Format.BLEFT)
            index += 1
            
        index = initial_index + len(neg_points) + len(pos_points) + 1
        section_number += 1
    
    writer.cell(29,index-1,"",Format.BTOP)
    
    
    COL_WIDTH = [ 0.12 for i in range(0,50) ]
    annexUtils.set_column(worksheet,COL_WIDTH)
    annexUtils.set_row_dict(worksheet,ROW_DICT)
    workbook.close()    



if __name__ == "__main__":
    generate(sys.argv[1],sys.argv[2])
