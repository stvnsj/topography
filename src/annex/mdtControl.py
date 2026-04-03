import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
from utils import str_to_flt
import sys
import reader as rd
import re
import annexUtils
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
from annexUtils import offset_range
from openpyxl   import load_workbook
from control.mdt import MDTControl
import level

ROW_DICT = {}

def generate (
        input1 = "",
        input2 = "",
        output_file = "/home/jstvns/axis/Anexos Formulas/mdtcontrol.xlsx",
        total_length = 1000,
) :
    
    workbook   = xlsxwriter.Workbook(output_file)
    worksheet  = workbook.add_worksheet("SHEET1")
    mdt_control = MDTControl(input1,input2)
    
    POINT_NUMBER = 1
    SEC_NUMBER = 1
    
    START_SECTION = SEC_NUMBER
    
    
    # worksheet.hide_gridlines(2) 
    # worksheet.set_portrait()
    # worksheet.set_page_view(2)
    # worksheet.set_paper(9)
    # worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)

    worksheet.hide_gridlines(2)  # 2 hides both the printed and visible gridlines
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(5)
    
    writer = Writer(workbook,worksheet)

    writer.range(2,19,2,6,"",Format.BORDER)
    writer.range(20,49,2,4,"VERIFICACIÓN DE PRECISIONES DEL MDT",
                 Format.SIZE(12),Format.BOLD, Format.BTOP,Format.BRIGHT, Format.CENTER, Format.VCENTER)
    writer.range(20,49,5,6,"FORMULARIO N° 2.309.307.A",
                 Format.SIZE(12),Format.BOLD, Format.BBOTTOM,Format.BRIGHT, Format.BOTTOM, Format.CENTER)

    
    
    writer.range(1,1,8,11,"",Format.BRIGHT)
    writer.range(50,50,8,11,"",Format.BLEFT)
    writer.range(2,49,7,7,"",Format.BBOTTOM)
    writer.range(2,49,12,12,"",Format.BTOP)
    
    writer.cell (2,8,"PROYECTO:",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.cell (2,9,"SECTOR:", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.cell (2,10,"TRAMO:", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.cell (2,11,"REALIZADO:",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.range(36,49,11,11,f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
    
    writer.cell(10,8, ": *")
    writer.cell(10,9, ": *")
    writer.cell(10,10,": *")
    writer.cell(10,11,": AUTOCONTROL TOPOGRÁFICO")
    
    
    
    index = 25
    ROW  = (index,index)
    
    
    dm_range = (2,7)
    n_range = offset_range(1,6,dm_range)
    side_range = offset_range(1,3,n_range)
    dist_range = offset_range(1,6,side_range)
    hc_range = offset_range(1,6,dist_range)
    hp_range = offset_range(1,6,hc_range)
    dif_range = offset_range(1,5,hp_range)
    absdif_range = offset_range(1,5,dif_range)
    good_range = offset_range(1,5,absdif_range)
    
    HEAD_FORMAT = (
        Format.VCENTER,
        Format.CENTER,
        Format.SIZE(10),
        Format.BORDER,
        Format.CENTER,
        Format.DOT
    )
    
    writer.range(*dm_range,index,index+1, "Perfil N°\n(Dm)", *HEAD_FORMAT)
    writer.range(*n_range,index,index+1,'Punto N°', *HEAD_FORMAT)
    writer.range(*side_range,index,index+1, "Lado", *HEAD_FORMAT)
    writer.range(*dist_range,index,index+1, "Distancia", *HEAD_FORMAT)
    writer.range(*hc_range,index,index+1, "Cota\nControl", *HEAD_FORMAT)
    writer.range(*hp_range,index,index+1,'Cota\nEstudio', *HEAD_FORMAT)
    writer.range(*dif_range,index,index+1,'Dif. (m)', *HEAD_FORMAT)
    writer.range(*absdif_range,index,index+1,'Dif. (m)\nValor Abs.', *HEAD_FORMAT)
    writer.range(*good_range,index,index+1,'Cumple\n(S/N)', *HEAD_FORMAT)
    
    
    index+=2
    DATA_FORMAT = (Format.CENTER,Format.SIZE(10),Format.BORDER,Format.NUM)

    
    for sec in mdt_control.section_list:
        writer.range(*dm_range, index, index + sec.point_length()-1, f'{SEC_NUMBER} ({sec.dm})' ,
                     Format.CENTER, Format.SIZE(10), Format.BORDER, Format.NUM, Format.VCENTER)
        SEC_NUMBER += 1
        for point in sec.point_list:
            COL = (index,index)
            
            writer.range(*n_range,    *COL, POINT_NUMBER, Format.CENTER,Format.SIZE(10),Format.BORDER)
            writer.range(*side_range, *COL, point.get_side(), *DATA_FORMAT)
            writer.range(*dist_range, *COL, point.distance, *DATA_FORMAT)
            writer.range(*hc_range,   *COL, point.ctrl_height, *DATA_FORMAT)
            writer.range(*hp_range,   *COL, point.proj_height, *DATA_FORMAT)
            writer.range(*dif_range,  *COL, point.delta(), *DATA_FORMAT)
            writer.range(*absdif_range, *COL, point.abs_delta(), *DATA_FORMAT)
            writer.range(*good_range, *COL, "SI" if point.is_within_tolerance() else "NO", *DATA_FORMAT)
            index+=1
            POINT_NUMBER += 1
 
 
    
    writer.range(2,49,index,index,"",Format.BTOP)
 
    
    writer.range(2,49,15,15,f'PERFIL {START_SECTION} AL PERFIL {SEC_NUMBER}',
                 Format.CENTER,Format.BOLD)
    
    
    ran1 = (2,15)
    ran2 = offset_range(3,5,ran1)
    ran3 = offset_range(5,19,ran2)
    ran4 = offset_range(3,5,ran3)
    
    FIELD_FORMAT = (Format.CENTER, Format.BBOTTOM, Format.SIZE(10))
    
    writer.range(ran1[0],ran2[1],17,17,'ANTECEDENTES GENERALES', Format.BOLD, Format.CENTER)
    writer.range(*ran1,18,18,'N° de Perfiles Contratados')
    writer.range(*ran1,19,19,'N° Total de Perfiles Levantados')
    writer.range(*ran1,20,20,'Cumple (S/N)')
    
    writer.range(*ran2,18,18,'*',*FIELD_FORMAT)
    writer.range(*ran2,19,19,mdt_control.section_length(),*FIELD_FORMAT)
    writer.range(*ran2,20,20,'*',*FIELD_FORMAT)
    
    writer.range(ran3[0],ran4[1],17,17,'RESULTADOS DE AUTOCONTROL', Format.BOLD, Format.CENTER)
    writer.range(*ran3,18,18,'Tolerancia en Cota (m)')
    writer.range(*ran3,19,19,'N° Puntos Controlados')
    writer.range(*ran3,20,20,'N° Puntos en Tolerancia')
    writer.range(*ran3,21,21,'% Puntos de Tolerancia')
    
    writer.range(*ran4,18,18,0.5,*FIELD_FORMAT)
    writer.range(*ran4,19,19,mdt_control.get_total_points(),*FIELD_FORMAT)
    writer.range(*ran4,20,20,mdt_control.get_good_points(),*FIELD_FORMAT)
    writer.range(*ran4,21,21,mdt_control.get_good_percent(),*FIELD_FORMAT)
    
    
    COL_WIDTH = [ 0.12 for i in range(0,50) ]
    
    annexUtils.set_column(worksheet,COL_WIDTH)
    workbook.close()
    



if __name__ == "__main__":
    generate(sys.argv[1],sys.argv[2])
