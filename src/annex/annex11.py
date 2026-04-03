import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
import sys
import reader as rd
import re
import annexUtils
import annex.scanner10
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
from openpyxl   import load_workbook

ROW_DICT = {
    0:0.1,
    6:0.1,
    12:0.1,
    10:0.1,
    16:0.13,
}

def contains_g(s):
    return 'g' in s or 'G' in s

def contains_t(s):
    return 't' in s or 'T' in s

def generate (input_file = 'anexos/anteproyecto/anexo10.xlsx', output_file = "test11.xlsx") :
    
    workbook = xlsxwriter.Workbook(output_file)

    worksheet = workbook.add_worksheet("Cotas PR")
    worksheetG = workbook.add_worksheet("Cotas G")
    worksheetT = workbook.add_worksheet("Cotas T")
    
    generatePR(workbook, worksheet, input_file)
    generateG(workbook, worksheetG, input_file)
    generateT(workbook, worksheetT, input_file)
    
    workbook.close()

##################################################################
#  __               _____ _____ _____ ___   _____  ____________  #
# /  |             /  __ \  _  |_   _/ _ \ /  ___| | ___ \ ___ \ #
# `| |    ______   | /  \/ | | | | |/ /_\ \\ `--.  | |_/ / |_/ / #
#  | |   |______|  | |   | | | | | ||  _  | `--. \ |  __/|    /  #
# _| |_            | \__/\ \_/ / | || | | |/\__/ / | |   | |\ \  #
# \___/             \____/\___/  \_/\_| |_/\____/  \_|   \_| \_| #
##################################################################

def generatePR (workbook,worksheet, input_file) :
    
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)
    
    writer = Writer(workbook,worksheet)
    
    wb = load_workbook(input_file)
    sheet_names = wb.sheetnames
    ws  = wb[sheet_names[0]]
    scanner = annex.scanner10.Scanner(ws)

    
    # FIXED CONTENT
    writer.merge(f"B2:D6","",Format.BORDER)
    writer.merge(
        f"E2:I5","COTAS DE PR",
        Format.BTOP,Format.BRIGHT,Format.SIZE(12),Format.BOLD,
        Format.CENTER, Format.VCENTER
    )
    
    writer.merge(
        f"E6:I6",
        "LÁMINA N° 2.903.3.I    FIGURA 2",
        Format.BBOTTOM,Format.BRIGHT,Format.SIZE(12),
        Format.BOLD,Format.CENTER
    )
    
    
    writer.write(f"J2","",Format.BLEFT)
    writer.merge(f"B7:I7","",{"bottom":1})
    writer.merge(f"A8:A12","",{"right":1})
    writer.merge(f"J8:J12","",{"left":1})
    writer.merge(f"B13:I13", "",{"top":1})
    
    writer.write(f"B8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"G12:I12",f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
    
  
    writer.merge(
        f'B14:C14',
        'PR',
        Format.BORDER,
        Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(
        f'D14:F14',
        'PROYECTO DEFINITIVO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(
        f'B15:B16',
        'DESDE',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'C15:C16',
        'HASTA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'D15:D16',
        'DESNIVEL\nIDA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'E15:E16',
        'DESNIVEL\nREGRESO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'F15:F16',
        'ERROR',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'G14:G16',
        'DESNIVEL\nPROMEDIO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'H14:H16',
        'COTA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'I14:I16',
        'ESTACION -\nPR',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(f'B17:I17','',Format.BORDER)
    


    curr_row = 18

    
 
    for block in scanner.blocks : 

        if curr_row == 18:
            writer.write(f'B{curr_row}', "",Format.CENTER,Format.SIZE(10),Format.BORDER)
            writer.write(f'C{curr_row}', block.start_point,Format.CENTER,Format.SIZE(10),Format.BORDER)
            writer.write(f'D{curr_row}', "",Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
            writer.write(f'E{curr_row}', "",Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
            writer.write(f'F{curr_row}', "",Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
            writer.write(f'G{curr_row}', "",Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
            writer.write(f'H{curr_row}', block.start_z,Format.CENTER,Format.SIZE(10),Format.BOLD,Format.BORDER, Format.NUM)
            writer.write(f'I{curr_row}', "",Format.CENTER,Format.SIZE(10),Format.BORDER)
            curr_row+=1
        

        writer.write(f'B{curr_row}', block.start_point,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'C{curr_row}', block.end_point,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'D{curr_row}', block.go_level,Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
        writer.write(f'E{curr_row}', block.come_level,Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
        writer.write(f'F{curr_row}', block.delta,Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
        writer.write(f'G{curr_row}', block.mean_level,Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
        writer.write(f'H{curr_row}', block.end_z,Format.CENTER,Format.SIZE(10),Format.BOLD,Format.BORDER, Format.NUM)
        writer.write(f'I{curr_row}', block.end_point,Format.CENTER,Format.SIZE(10),Format.BORDER)
        
        
        curr_row += 1


    
    writer.merge(f'B{curr_row}:I{curr_row}', "", Format.BTOP)
    
    formatter = Formatter(worksheet)
    col_width = {
        0:0.13, #A
        1:0.86, #B
        2:0.79, #C
        3:0.79, #D
        4:0.79, #E
        5:0.79, #F
        6:0.85, #G
        7:0.79, #H
        8:0.85, #I
        9:0.13, #J
    }
    annexUtils.set_row_dict(worksheet,ROW_DICT)
    # worksheet.autofit()
    formatter.set_cols(col_width)
    # formatter.set_rows({0:2,1:2,2:2, 4:2})
    
    
############################################################
#  __               _____ _____ _____ ___   _____   _____  #
# /  |             /  __ \  _  |_   _/ _ \ /  ___| |  __ \ #
# `| |    ______   | /  \/ | | | | |/ /_\ \\ `--.  | |  \/ #
#  | |   |______|  | |   | | | | | ||  _  | `--. \ | | __  #
# _| |_            | \__/\ \_/ / | || | | |/\__/ / | |_\ \ #
# \___/             \____/\___/  \_/\_| |_/\____/   \____/ #
############################################################
def generateG (workbook,worksheet,input_file) :
    
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)
    
    writer = Writer(workbook,worksheet)
    
    wb = load_workbook(input_file)
    sheet_names = wb.sheetnames
    ws  = wb[sheet_names[1]]
    scanner = annex.scanner10.Scanner(ws)
    
    # FIXED CONTENT
    writer.merge(f"B2:D6","",Format.BORDER)
    writer.merge(
        f"E2:I5","COTAS DE PR",
        Format.BTOP,Format.BRIGHT,Format.SIZE(12),Format.BOLD,
        Format.CENTER, Format.VCENTER
    )
    
    writer.merge(
        f"E6:I6",
        "LÁMINA N° 2.903.3.I    FIGURA 2",
        Format.BBOTTOM,Format.BRIGHT,Format.SIZE(12),
        Format.BOLD,Format.CENTER
    )
    
    
    writer.write(f"J2","",Format.BLEFT)
    writer.merge(f"B7:I7","",{"bottom":1})
    writer.merge(f"A8:A12","",{"right":1})
    writer.merge(f"J8:J12","",{"left":1})
    writer.merge(f"B13:I13", "",{"top":1})
    
    writer.write(f"B8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"G12:I12",f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
  
    writer.merge(
        f'B14:C14',
        'PR',
        Format.BORDER,
        Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(
        f'D14:F14',
        'PROYECTO DEFINITIVO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(
        f'B15:B16',
        'DESDE',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'C15:C16',
        'HASTA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'D15:D16',
        'DESNIVEL\nIDA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'E15:E16',
        'DESNIVEL\nREGRESO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'F15:F16',
        'ERROR',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'G14:G16',
        'DESNIVEL\nPROMEDIO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'H14:H16',
        'COTA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'I14:I16',
        'ESTACION -\nPR',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(f'B17:I17','',Format.BORDER)

    curr_row = 19

    for block in scanner.blocks:
        if contains_g(block.end_point):
            writer.write(f'B{curr_row-1}', '', Format.BORDER)
            writer.write(f'C{curr_row-1}', '', Format.BORDER)
            writer.write(f'D{curr_row-1}', '', Format.BORDER)
            writer.write(f'E{curr_row-1}', '', Format.BORDER)
            writer.write(f'F{curr_row-1}', '', Format.BORDER)
            writer.write(f'G{curr_row-1}', '', Format.BORDER)
            writer.write(f'H{curr_row-1}', block.start_z,Format.CENTER,Format.SIZE(10),Format.BORDER)
            writer.write(f'I{curr_row-1}', block.start_point,Format.CENTER,Format.SIZE(10),Format.BORDER)        
            
            writer.write(f'B{curr_row}', block.start_point,Format.CENTER,Format.SIZE(10),Format.BORDER)
            writer.write(f'C{curr_row}', block.end_point,Format.CENTER,Format.SIZE(10),Format.BORDER)
            writer.write(f'D{curr_row}', block.go_level,Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
            writer.write(f'E{curr_row}', block.come_level,Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
            writer.write(f'F{curr_row}', block.delta,Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
            writer.write(f'G{curr_row}', block.mean_level,Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
            writer.write(f'H{curr_row}', block.end_z,Format.CENTER,Format.SIZE(10),Format.BOLD,Format.BORDER, Format.NUM)
            writer.write(f'I{curr_row}', block.end_point,Format.CENTER,Format.SIZE(10),Format.BORDER)
            
            writer.write(f'B{curr_row+1}', '', Format.BORDER)
            writer.write(f'C{curr_row+1}', '', Format.BORDER)
            writer.write(f'D{curr_row+1}', '', Format.BORDER)
            writer.write(f'E{curr_row+1}', '', Format.BORDER)
            writer.write(f'F{curr_row+1}', '', Format.BORDER)
            writer.write(f'G{curr_row+1}', '', Format.BORDER)
            writer.write(f'H{curr_row+1}', '', Format.BORDER)
            writer.write(f'I{curr_row+1}', '', Format.BORDER) 
            
            curr_row += 3
    
    
    writer.merge(f'B{curr_row-1}:I{curr_row-1}', "", Format.BTOP)
    
    formatter = Formatter(worksheet)
    col_width = {
        0:0.13, #A
        1:0.79, #B
        2:0.79, #C
        3:0.79, #D
        4:0.79, #E
        5:0.79, #F
        6:0.85, #G
        7:0.79, #H
        8:0.85, #I
        9:0.13, #J
    }
    annexUtils.set_row_dict(worksheet,ROW_DICT)
    formatter.set_cols(col_width)


############################################################
#  __               _____ _____ _____ ___   _____   _____  #
# /  |             /  __ \  _  |_   _/ _ \ /  ___| |_   _| #
# `| |    ______   | /  \/ | | | | |/ /_\ \\ `--.    | |   #
#  | |   |______|  | |   | | | | | ||  _  | `--. \   | |   #
# _| |_            | \__/\ \_/ / | || | | |/\__/ /   | |   #
# \___/             \____/\___/  \_/\_| |_/\____/    \_/   #
############################################################
def generateT(workbook,worksheet,input_file):
     
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)
    
    writer = Writer(workbook,worksheet)
    
    wb = load_workbook(input_file)
    sheet_names = wb.sheetnames
    ws  = wb[sheet_names[2]]
    scanner = annex.scanner10.Scanner(ws)
    
    # FIXED CONTENT
    writer.merge(f"B2:D6","",Format.BORDER)
    writer.merge(
        f"E2:I5","COTAS DE PR",
        Format.BTOP,Format.BRIGHT,Format.SIZE(12),Format.BOLD,
        Format.CENTER, Format.VCENTER
    )
    
    writer.merge(
        f"E6:I6",
        "LÁMINA N° 2.903.3.I    FIGURA 2",
        Format.BBOTTOM,Format.BRIGHT,Format.SIZE(12),
        Format.BOLD,Format.CENTER
    )
    
    
    writer.write(f"J2","",Format.BLEFT)
    writer.merge(f"B7:I7","",{"bottom":1})
    writer.merge(f"A8:A12","",{"right":1})
    writer.merge(f"J8:J12","",{"left":1})
    writer.merge(f"B13:I13", "",{"top":1})
    
    writer.write(f"B8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"G12:I12",f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
  
    writer.merge(
        f'B14:C14',
        'PR',
        Format.BORDER,
        Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(
        f'D14:F14',
        'PROYECTO DEFINITIVO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(
        f'B15:B16',
        'DESDE',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'C15:C16',
        'HASTA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'D15:D16',
        'DESNIVEL\nIDA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'E15:E16',
        'DESNIVEL\nREGRESO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'F15:F16',
        'ERROR',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'G14:G16',
        'DESNIVEL\nPROMEDIO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'H14:H16',
        'COTA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'I14:I16',
        'ESTACION -\nPR',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(f'B17:I17','',Format.BORDER)

    curr_row = 19

    for block in scanner.blocks:
        if contains_t(block.end_point):
            writer.write(f'B{curr_row-1}', '', Format.BORDER)
            writer.write(f'C{curr_row-1}', '', Format.BORDER)
            writer.write(f'D{curr_row-1}', '', Format.BORDER)
            writer.write(f'E{curr_row-1}', '', Format.BORDER)
            writer.write(f'F{curr_row-1}', '', Format.BORDER)
            writer.write(f'G{curr_row-1}', '', Format.BORDER)
            writer.write(f'H{curr_row-1}', block.start_z,Format.CENTER,Format.SIZE(10),Format.BORDER)
            writer.write(f'I{curr_row-1}', block.start_point,Format.CENTER,Format.SIZE(10),Format.BORDER)        
            
            writer.write(f'B{curr_row}', block.start_point,Format.CENTER,Format.SIZE(10),Format.BORDER)
            writer.write(f'C{curr_row}', block.end_point,Format.CENTER,Format.SIZE(10),Format.BORDER)
            writer.write(f'D{curr_row}', block.go_level,Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
            writer.write(f'E{curr_row}', block.come_level,Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
            writer.write(f'F{curr_row}', block.delta,Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
            writer.write(f'G{curr_row}', block.mean_level,Format.CENTER,Format.SIZE(10),Format.BORDER, Format.NUM)
            writer.write(f'H{curr_row}', block.end_z,Format.CENTER,Format.SIZE(10),Format.BOLD,Format.BORDER, Format.NUM)
            writer.write(f'I{curr_row}', block.end_point,Format.CENTER,Format.SIZE(10),Format.BORDER)
            
            writer.write(f'B{curr_row+1}', '', Format.BORDER)
            writer.write(f'C{curr_row+1}', '', Format.BORDER)
            writer.write(f'D{curr_row+1}', '', Format.BORDER)
            writer.write(f'E{curr_row+1}', '', Format.BORDER)
            writer.write(f'F{curr_row+1}', '', Format.BORDER)
            writer.write(f'G{curr_row+1}', '', Format.BORDER)
            writer.write(f'H{curr_row+1}', '', Format.BORDER)
            writer.write(f'I{curr_row+1}', '', Format.BORDER) 
            
            curr_row += 3
    
    
    writer.merge(f'B{curr_row-1}:I{curr_row-1}', "", Format.BTOP)
    
    formatter = Formatter(worksheet)
    col_width = {
        0:0.13, #A
        1:0.79, #B
        2:0.79, #C
        3:0.79, #D
        4:0.79, #E
        5:0.79, #F
        6:0.85, #G
        7:0.79, #H
        8:0.85, #I
        9:0.13, #J
    }
    annexUtils.set_row_dict(worksheet,ROW_DICT)
    formatter.set_cols(col_width)





if __name__ == "__main__":
    generate()
    
    
